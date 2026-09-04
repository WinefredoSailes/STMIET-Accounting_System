"""Bulk-import the customer master (BUILD-PLAN Phase 10 master-data migration).

Usage:
    py manage.py import_customers --file excel-files/CUSTOMERS.xlsx
    py manage.py import_customers --file excel-files/CUSTOMERS.csv

Expected columns (header row may be present; the importer locates columns by
header name, matching the pattern of import_coa / import_fixed_assets):

    CODE | NAME | SEGMENT | GROUP | PRICING TIER | TIN | ADDRESS | CONTACT | NOTES

  - SEGMENT/SECTION    : "DHPP", "DMIE" or "OPS" (required)
  - GROUP              : fuel | equipment | ops (defaults to fuel)
  - PRICING TIER       : regular | patron | volume (defaults to regular)
  - Everything else optional.

The importer is idempotent: a row whose CODE already exists is updated
(in-place) rather than duplicated — re-running is safe.

If no --file is given, the command falls back to any file matching one of
the candidate names under /excel-files.
"""

import csv
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.ar.models import Customer, CustomerGroup, PricingTier
from apps.foundation.models import Company, Segment

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

HEADER_MAP = {
    "code": ("CODE", "CUSTOMER CODE", "CUST CODE", "ACCT #"),
    "name": ("NAME", "CUSTOMER NAME", "OUTLET NAME", "CLIENT"),
    "segment": ("SEGMENT", "SECTION", "DEPARTMENT"),
    "group": ("GROUP", "CATEGORY"),
    "pricing_tier": ("PRICING TIER", "TIER", "PRICE TIER"),
    "tin": ("TIN",),
    "address": ("ADDRESS",),
    "contact": ("CONTACT", "CONTACT NO", "PHONE", "MOBILE", "TEL NO"),
    "notes": ("NOTES", "REMARKS"),
}

SEGMENT_ALIASES = {
    "DHPP": "DHPP", "DIESEL & HEAVY PARTS PROCUREMENT": "DHPP", "DIESEL": "DHPP",
    "DMIE": "DMIE", "DIESEL MACHINERY & INDUSTRIAL EQUIPMENT": "DMIE",
    "OPS": "OPS", "OPERATIONS": "OPS", "OPERATIONS / SERVICES": "OPS",
}

GROUP_ALIASES = {
    "FUEL": CustomerGroup.FUEL, "DIESEL": CustomerGroup.FUEL,
    "EQUIPMENT": CustomerGroup.EQUIPMENT, "MACHINERY": CustomerGroup.EQUIPMENT,
    "OPS": CustomerGroup.OPS, "OPERATIONS": CustomerGroup.OPS, "SERVICES": CustomerGroup.OPS,
}

TIER_ALIASES = {
    "REGULAR": PricingTier.REGULAR,
    "PATRON": PricingTier.PATRON,
    "VOLUME": PricingTier.VOLUME,
}


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Bulk-import the customer master from CSV or XLSX (idempotent)."

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file", default=None, help="CSV/XLSX path")
        parser.add_argument("--company", dest="company", default="STMIET")

    def _load_rows(self, file_path):
        """Yield (row_dict, row_label) for each data row."""
        if file_path.suffix.lower() == ".csv":
            with open(file_path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.reader(fh)
                header = [h.strip().upper() for h in next(reader, [])]
                for i, row in enumerate(reader, start=2):
                    if not any(row):
                        continue
                    yield self._map_row(row, header), f"CSV row {i}"
        else:
            if openpyxl is None:
                raise CommandError("openpyxl required for xlsx input; pip install openpyxl")
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header = None
                header_idx = 0
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if not row or not any(row):
                        continue
                    upper = [str(v or "").strip().upper() for v in row]
                    if header is None and any("NAME" in h or "CODE" in h for h in upper):
                        header = upper
                        header_idx = i
                        break
                if header is None:
                    self.stdout.write(self.style.WARNING(
                        f"skip sheet {sheet_name}: no recognizable header row"
                    ))
                    continue
                for i, row in enumerate(ws.iter_rows(min_row=header_idx + 1, values_only=True), start=header_idx + 1):
                    if not row or not any(row):
                        continue
                    yield self._map_row(row, header), f"{sheet_name} row {i}"

    def _map_row(self, row, header):
        """Build {field: value} from a column position map."""
        idx = {}
        for field, names in HEADER_MAP.items():
            for n in names:
                if n in header:
                    idx[field] = header.index(n)
                    break
        out = {}
        for field, pos in idx.items():
            if pos < len(row):
                out[field] = _clean(row[pos])
        return out

    def _resolve_segment(self, raw):
        key = raw.upper().strip()
        segment_code = SEGMENT_ALIASES.get(key)
        if not segment_code:
            # Match partial segment text (e.g. "DHPP" inside a longer cell).
            for alias, code in SEGMENT_ALIASES.items():
                if alias and alias in key:
                    segment_code = code
                    break
        if not segment_code:
            raise ValueError(f"Unknown segment: {raw!r}")
        segment = Segment.objects.filter(code=segment_code).first()
        if segment is None:
            raise CommandError(
                f"Segment '{segment_code}' not found. Run import_coa first."
            )
        return segment

    def _resolve_group(self, raw):
        if not raw:
            return CustomerGroup.FUEL
        key = raw.upper().strip()
        return GROUP_ALIASES.get(key, CustomerGroup.FUEL)

    def _resolve_tier(self, raw):
        if not raw:
            return PricingTier.REGULAR
        return TIER_ALIASES.get(raw.upper().strip(), PricingTier.REGULAR)

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = Path(options["file"]) if options["file"] else None
        if file_path is None or not file_path.exists():
            repo_root = Path(__file__).resolve().parents[5]
            for cand in (
                repo_root / "excel-files" / "CUSTOMERS.xlsx",
                repo_root / "excel-files" / "CUSTOMERS.csv",
                Path.cwd() / "excel-files" / "CUSTOMERS.xlsx",
            ):
                if cand.exists():
                    file_path = cand
                    break
        if file_path is None or not file_path.exists():
            raise CommandError(
                "Customer master not found. Pass --file, or place CUSTOMERS.csv"
                "/.xlsx under /excel-files."
            )

        company = Company.objects.filter(code=options["company"]).first()
        if company is None:
            raise CommandError(f"Company '{options['company']}' not found. Run import_coa first.")

        created = updated = skipped = 0
        for row, label in self._load_rows(file_path):
            code = row.get("code", "")
            name = row.get("name", "")
            if not code and not name:
                skipped += 1
                continue
            try:
                segment = self._resolve_segment(row.get("segment", "ALL"))
                if not code:
                    base = re.sub(r"[^A-Z0-9]+", "-", name.upper())[:16].strip("-")
                    code = f"{base}-{derived_seq()}"
                group = self._resolve_group(row.get("group", ""))
                tier = self._resolve_tier(row.get("pricing_tier", ""))
                _, was_created = Customer.objects.update_or_create(
                    code=code,
                    defaults={
                        "name": name or code,
                        "group": group,
                        "segment": segment,
                        "pricing_tier": tier,
                        "tin": row.get("tin", ""),
                        "address": row.get("address", ""),
                        "contact_no": row.get("contact", ""),
                        "notes": row.get("notes", ""),
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except (ValueError, CommandError) as exc:
                self.stdout.write(self.style.ERROR(f"  {label}: {exc}"))
                skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f"import_customers: {created} created, {updated} updated, {skipped} skipped."
        ))


def derived_seq():
    """Small counter for auto-generated customer codes (stateless per run)."""
    n = getattr(derived_seq, "_n", 0) + 1
    derived_seq._n = n
    return f"{n:03d}"
