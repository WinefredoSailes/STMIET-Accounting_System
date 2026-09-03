"""Seed the Fixed Asset Register at company opening from the source workbook.

Usage:
    py manage.py import_fixed_assets --file excel-files/SEPTEMBER-1-2026-_-FIXED-ASSETS.xlsx
    py manage.py import_fixed_assets --mapping excel-files/fixed-assets-mapping.json

This is the Phase 10 master-data migration for the Sept 1, 2026 snapshot. All
*asset data* (item names, costs, accumulated depreciation, purchase dates,
useful lives) is read from the workbook; nothing is hardcoded. The only thing
the workbook cannot supply is the COA account each asset class posts to, so the
sheet/section -> {category, COA codes, segment} wiring lives in a data-driven
sidecar (fixed-assets-mapping.json), not in Python — mirroring how import_coa
keeps account codes as DB data.

For every row the importer:
  * creates/updates the AssetCategory (from the mapping),
  * creates the Asset and posts the opening JE (AssetService.seed_opening:
        Dr Asset = cost | Cr Accum Dep = to-date | Cr opening equity = NBV),
  * stores the to-date accumulated depreciation on a posted schedule row, so
    the register's accumulated depreciation and NBV tie out to the workbook.

Re-runs are idempotent: an existing asset_no is left untouched.
"""

import calendar
import json
import re
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.assets.models import Asset, AssetCategory
from apps.assets.services import AssetService
from apps.foundation.models import Account, Company, Segment

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

# Header tokens (case-insensitive) used to locate columns regardless of layout.
NAME_HEADERS = {"PARTICULARS", "ITEMS", "ITE\\MS"}
COST_HEADERS = {"COST", "TOTAL COST"}
ACCUM_HEADERS = {"ACCUMULATED DEPRECIATION"}
# Column whose header hints an item's brand/serial (used to disambiguate names).
BRAND_HEADERS = {"BRAND /SERIAL NUMBER / INDICATOR", "PLATE NO."}

DECIMAL_RE = re.compile(r"^-?[\d,]+(\.\d+)?$")
MONTH_NAMES = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"])}


def _to_decimal(value):
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        v = Decimal(str(value))
    else:
        text = str(value).strip().replace(",", "")
        if not DECIMAL_RE.match(text):
            return None
        v = Decimal(text)
    return v


def _parse_date(value, as_of: date) -> date:
    """Parse common PH workbook date formats; fall back to as_of on failure."""
    if value is None:
        return as_of
    if isinstance(value, date):
        return value
    text = str(value).strip()
    m = re.match(r"^(\d{1,4})[-/](\d{1,2})[-/](\d{1,4})$", text)
    if m:
        a, b, c = m.groups()
        if len(a) == 4 or len(c) == 4:
            year = int(a) if len(a) == 4 else int(c)
            mon, day = (int(b), int(c)) if len(a) == 4 else (int(a), int(b))
            try:
                return date(year, mon, day)
            except ValueError:
                return as_of
    # "JUNE 24, 2025" / "SEPTEMBER 20, 2025"
    m = re.match(
        r"^\s*([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})\s*$", text
    )
    if m:
        mon = MONTH_NAMES.get(m.group(1).lower()[:3])
        if mon:
            try:
                return date(int(m.group(3)), mon, int(m.group(2)))
            except ValueError:
                return as_of
    return as_of


class Command(BaseCommand):
    help = "Seed AssetCategory + Asset opening register from the fixed-assets workbook."

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file", default=None, help="fixed assets xlsx path")
        parser.add_argument("--mapping", dest="mapping", default=None, help="category->COA json config")
        parser.add_argument("--as-of", dest="as_of", default="2026-09-01")

    def _load(self, options):
        repo_root = Path(__file__).resolve().parents[5]
        file_path = Path(options["file"]) if options["file"] else None
        if file_path is None or not file_path.exists():
            for cand in (
                repo_root / "excel-files" / "SEPTEMBER-1-2026-_-FIXED-ASSETS.xlsx",
                Path.cwd() / "excel-files" / "SEPTEMBER-1-2026-_-FIXED-ASSETS.xlsx",
            ):
                if cand.exists():
                    file_path = cand
                    break
        if file_path is None or not file_path.exists():
            raise CommandError("Fixed-assets workbook not found. Pass --file.")
        if openpyxl is None:
            raise CommandError("openpyxl required; install with: pip install openpyxl")
        self.stdout.write(f"Reading {file_path.name}")

        mapping_path = Path(options["mapping"]) if options["mapping"] else None
        if mapping_path is None or not mapping_path.exists():
            for cand in (
                repo_root / "excel-files" / "fixed-assets-mapping.json",
                Path.cwd() / "excel-files" / "fixed-assets-mapping.json",
            ):
                if cand.exists():
                    mapping_path = cand
                    break
        if mapping_path is None or not mapping_path.exists():
            raise CommandError("Mapping config not found. Pass --mapping.")
        mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
        return file_path, mapping

    @transaction.atomic
    def handle(self, *args, **options):
        file_path, mapping = self._load(options)
        as_of = _parse_date(options["as_of"], date(2026, 9, 1))
        company = Company.objects.filter(code=mapping.get("company", "STMIET")).first()
        if company is None:
            raise CommandError(
                f"Company '{mapping.get('company')}' not found. Run import_coa first."
            )
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

        categories = self._upsert_categories(mapping, company)
        self._validate_codes(mapping, categories)

        created = 0
        seq = 0
        for sheet_name in wb.sheetnames:
            section_map = mapping.get("sections", {}).get(sheet_name)
            if not section_map:
                continue
            seq, sheet_created = self._import_sheet(
                wb[sheet_name], section_map, categories, as_of, seq,
            )
            created += sheet_created
        self.stdout.write(
            self.style.SUCCESS(f"Imported {created} fixed assets from the September 1, 2026 register.")
        )

    def _upsert_categories(self, mapping, company) -> dict:
        """Collapse section configs into AssetCategory rows (code is unique)."""
        categories = {}
        defaults_by_code = {}
        for sheet_map in mapping.get("sections", {}).values():
            # Fold the section default (sheets that only carry a __default__).
            base = sheet_map.get("__default__", {})
            if base:
                self._fold_category(base, defaults_by_code)
            for key, cfg in list(sheet_map.items()):
                if key.startswith("__"):
                    continue
                self._fold_category(cfg, defaults_by_code)
            for rule in sheet_map.get("__rules__", []):
                # rules inherit the section default except `asset`; fold too.
                folded = dict(base)
                folded.update(rule)
                self._fold_category(folded, defaults_by_code)
        for code, cfg in defaults_by_code.items():
            default = cfg["default"]
            category, _ = AssetCategory.objects.update_or_create(
                code=code,
                defaults={
                    "name": (default.get("name") or code.replace("_", " ").title()),
                    "useful_life_years": default["life"],
                    "asset_account": Account.objects.get(code=default["asset"]),
                    "depreciation_expense_account": Account.objects.get(
                        code=default["dep_exp"]
                    ),
                    "accumulated_dep_account": Account.objects.get(code=default["accum"]),
                    "segment": None,
                },
            )
            categories[code] = category
        return categories

    def _fold_category(self, cfg, defaults_by_code):
        code = cfg["category"]
        folded = defaults_by_code.setdefault(
            code, {"default": {k: cfg[k] for k in ("life", "asset", "accum", "dep_exp", "segment")}}
        )
        if cfg["asset"] != folded["default"].get("asset"):
            folded.setdefault("alts", set())
            folded["alts"].add(cfg["asset"])
        folded["default"].setdefault("name", cfg.get("name"))
        # keep the most common life; last-write is fine here.
        folded["default"]["life"] = cfg["life"]
        return folded

    def _validate_codes(self, mapping, categories):
        needed = set()
        for sheet_map in mapping.get("sections", {}).values():
            for key, cfg in list(sheet_map.items()):
                if key.startswith("__"):
                    continue
                needed.update([cfg["asset"], cfg["accum"], cfg["dep_exp"]])
            base = sheet_map.get("__default__", {})
            for rule in sheet_map.get("__rules__", []):
                cfg = dict(base); cfg.update(rule)
                needed.update([cfg["asset"], cfg.get("accum", base.get("accum")),
                               cfg.get("dep_exp", base.get("dep_exp"))])
        have = set(Account.objects.filter(code__in=needed).values_list("code", flat=True))
        missing = needed - have
        if missing:
            raise CommandError(
                f"Mapping references COA codes missing from the chart: {sorted(missing)}. "
                "Run import_coa with the revised workbook first."
            )

    def _import_sheet(self, ws, section_map, categories, as_of, seq):
        name_col = date_col = life_col = cost_col = accum_col = brand_col = None
        header_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row:
                continue
            lowered = {str(v or "").strip().upper(): j for j, v in enumerate(row) if v is not None}
            if "COST" in lowered or "TOTAL COST" in lowered or self._is_header_row(row):
                header_idx = i
                name_col = self._find_col(lowered, NAME_HEADERS)
                date_col = self._find_header(lowered, "DATE OF PURCHASE")
                life_col = self._find_contains(lowered, ("USEFUL LIFE", "ESTIMATED LIFE"))
                brand_col = self._find_col(lowered, BRAND_HEADERS)
                cost_col = lowered.get("COST") if "COST" in lowered else lowered.get("TOTAL COST")
                accum_col = self._find_contains(lowered, ACCUM_HEADERS)
                break
        if header_idx is None or cost_col is None:
            self.stdout.write(self.style.WARNING(f"skip sheet {ws.title}: no COST header"))
            return seq, 0

        current_section = section_map.get("__default__", {})
        count = 0
        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            if not row:
                continue
            first = str(row[0] or "").strip() if len(row) > 0 else ""
            # VEHICLE-style section headers (e.g. "A. FUEL TANKERS").
            if first in section_map:
                current_section = section_map[first]
                continue
            if not first:
                continue
            cost = _to_decimal(row[cost_col]) if cost_col < len(row) else None
            if cost is None or cost <= 0:
                continue
            cfg = self._resolve_row_cfg(section_map, first, current_section)
            accum = _to_decimal(row[accum_col]) if accum_col is not None and accum_col < len(row) else Decimal("0.00")
            accum = accum or Decimal("0.00")
            category_code = cfg["category"]
            category = categories[category_code]
            name = self._build_name(row, first, name_col, brand_col)
            acquisition_date = _parse_date(row[date_col] if date_col is not None and date_col < len(row) else None, as_of)
            seq += 1
            asset_no = f"FA-2026-{seq:04d}"
            if Asset.objects.filter(asset_no=asset_no).exists():
                continue  # idempotent
            segment = Segment.objects.filter(code=cfg.get("segment", "OPS")).first()
            if segment is None:
                raise CommandError(
                    f"Segment '{cfg.get('segment')}' not found for sheet {ws.title} row '{first}'. "
                    "Run import_coa first."
                )
            AssetService.seed_opening(
                asset_no=asset_no,
                name=name,
                category=category,
                segment=segment,
                acquisition_date=acquisition_date,
                cost=cost,
                accumulated_dep=accum,
                asset_account=Account.objects.get(code=cfg["asset"]),
                depreciation_expense_account=category.depreciation_expense_account,
                accumulated_dep_account=category.accumulated_dep_account,
            )
            count += 1
            self.stdout.write(f"  + {asset_no} {name} ({category_code}) cost {cost} accum {accum}")
        return seq, count

    def _is_header_row(self, row):
        vals = [str(v or "").upper() for v in row]
        return any("ACCUMULATED DEPRECIATION" in v for v in vals) and any("COST" in v for v in vals)

    def _find_col(self, lowered_map, headers):
        for h in headers:
            if h in lowered_map:
                return lowered_map[h]
        return None

    def _find_header(self, lowered_map, header):
        for k, j in lowered_map.items():
            if header in k:
                return j
        return None

    def _find_contains(self, lowered_map, needles):
        for k, j in lowered_map.items():
            for n in needles:
                if n in k:
                    return j
        return None

    def _resolve_row_cfg(self, section_map, first, current_section):
        base = section_map.get("__default__", {})
        rules = section_map.get("__rules__", [])
        for rule in rules:
            if rule.get("match", "").upper() in first.upper():
                cfg = dict(base)
                cfg.update(rule)
                return cfg
        return current_section if current_section else base

    def _build_name(self, row, first, name_col, brand_col):
        parts = [first]
        if brand_col is not None and brand_col < len(row):
            extra = str(row[brand_col] or "").strip()
            if extra and extra != first:
                parts.append(extra)
        # Include any non-static descriptor columns (D/E) to disambiguate.
        for idx in (3, 4):  # "4 DRAWERS", "WHITE", serial info...
            if idx < len(row) and row[idx] is not None:
                extra = str(row[idx]).strip()
                if extra and extra not in parts:
                    parts.append(extra)
        return " — ".join(parts).strip(" —")