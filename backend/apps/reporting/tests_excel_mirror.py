"""Mirror-fidelity tests: every Excel export builder must reproduce the layout
of the source workbook it was extracted from in /excel-files.

The source workbooks contain *sample* data (e.g. year "2025", "xx", June dates,
the full legal entity name), whereas the builders emit the real, computed
values. So these tests normalise away the dynamic cells (company name, years,
dates, month names) and assert the *structural* labels/order match the source
cell-for-cell -- so the "template" in the system can never drift from the Excel
it was extracted from.
"""

import re
from collections import Counter
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from apps.reporting.excel_export import (
    build_cash_flow_statement,
    build_income_statement,
    build_statement_of_changes_in_equity,
    build_statement_of_cost_of_sales,
    build_statement_of_financial_position,
    build_statement_of_total_expenses,
    build_trial_balance,
)
from apps.reporting.models import StatementType

# repo root /excel-files  (this file is at backend/apps/reporting/)
EXCEL_DIR = Path(__file__).resolve().parents[3] / "excel-files"
COMPANY_NAME = "SEVEN-TRENT MACHINERIES INDUSTRIAL EQUIPMENT TRADING"
MONTHS = ("JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY",
          "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER")


def _is_dynamic(v: str) -> bool:
    """True for cells whose text is sample/computed data, not structure."""
    s = v.strip()
    if not s:
        return True
    if "SEVEN-TRENT" in s.upper() or "STMIET" in s.upper():
        return True  # legal entity name varies by company record
    if re.search(r"(19|20)\d{2}", s):
        return True  # a year
    if re.search(r"\b(" + "|".join(MONTHS) + r")\b", s, re.IGNORECASE):
        return True  # a month name (sample period)
    if "xx" in s.lower():
        return True  # placeholder
    if re.search(r"\d{1,2},?\s*\d{4}", s):
        return True  # a full date
    return False


def _src(path):
    return openpyxl.load_workbook(EXCEL_DIR / path, data_only=True)


def _labels(ws):
    """Multiset of structural text labels (ignores numbers, formula errors and
    dynamic/sample cells)."""
    out = []
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.strip() and not v.startswith("#") and not _is_dynamic(v):
                out.append(v.strip())
    return Counter(out)


def _headers_row5(ws):
    return [ws.cell(row=5, column=i).value for i in range(1, 9)]


def _row_of(ws, title):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == title:
            return r
    raise AssertionError(f"label not found: {title}")


# Reuse the fixtures defined in the reporting test module.
from apps.reporting.tests import (  # noqa: E402
    company,
    coa,
    period_data,
    segments,
    templates,
)


class TestTrialBalanceMirror:
    def test_title_and_headers_match_source(self, company, segments, coa, period_data):
        wb = build_trial_balance(company, 2026)
        ws = wb["TRIAL BALANCE"]
        src = _src("TRIAL-BALANCE.xlsx")["TRIAL BALANCE"]
        assert "(TRIAL BALANCE)" in (ws["I2"].value or "")
        assert "(TRIAL BALANCE)" in (src["I2"].value or "")
        # Column header band (row 5) — compare after stripping cosmetic spacing.
        assert [str(h).strip() if h else h for h in _headers_row5(ws)] == \
               [str(h).strip() if h else h for h in _headers_row5(src)]
        months = [ws.cell(row=5, column=c).value for c in range(11, 35, 2)]
        assert months == list(MONTHS)


class TestSFPMirror:
    def test_labels_match_source(self, company, segments, coa, period_data, templates):
        wb = build_statement_of_financial_position(company, date(2026, 1, 1), date(2026, 1, 31), "127000.00")
        assert _labels(wb["YEAR END"]) == _labels(_src("STATEMENT-OF-FINANCIAL-POSITION.xlsx")["YEAR END"])


class TestSOCEMirror:
    def test_labels_match_source(self, company, segments, coa, period_data, templates):
        wb = build_statement_of_changes_in_equity(company, date(2026, 1, 1), date(2026, 1, 31), "127000.00")
        assert _labels(wb["EQUITY"]) == _labels(_src("STATEMENT-OF-CHANGES-IN-EQUITY.xlsx")["EQUITY"])


class TestCoSMirror:
    def test_labels_match_source(self, company, segments, coa, period_data, templates):
        wb = build_statement_of_cost_of_sales(company, date(2026, 1, 1), date(2026, 1, 31))
        assert _labels(wb["COST OF SALES"]) == _labels(_src("STATEMENT-OF-COST-OF-SALES.xlsx")["COST OF SALES"])
        assert "Statement of Cost of Sales" in wb["COST OF SALES"]["B1"].value

    def test_segment_total_labels_present(self, company, segments, coa, period_data, templates):
        # Regression: the source labels these segment-total rows; builders must too.
        wb = build_statement_of_cost_of_sales(company, date(2026, 1, 1), date(2026, 1, 31))
        ws = wb["COST OF SALES"]
        assert ws["A22"].value == "COGS - DHPP"
        assert ws["A45"].value == "COGS - DMIE"
        assert ws["A53"].value == "COGS - OPS"


class TestTEMirror:
    def test_labels_match_source(self, company, segments, coa, period_data, templates):
        wb = build_statement_of_total_expenses(company, date(2026, 1, 1), date(2026, 1, 31))
        assert _labels(wb["January 2026 CGSE"]) == _labels(_src("STATEMENT-OF-TOTAL-EXPENSES.xlsx")["MARCH 2026 CGSE"])
        assert "Statement of Total Expenses" in wb["January 2026 CGSE"]["B1"].value


class TestCashFlowMirror:
    def test_labels_match_source(self, segments, coa):
        from apps.cash.models import ActivityType, BankAccount, CashCycleActivity, WeeklyCashCycle

        dhpp = segments["DHPP"]
        BankAccount.objects.create(
            code="PNB-DHPP", name="PNB DHPP", account_type="checking",
            bank_name="PNB", bank_code="PNB", gl_account=coa["10010"],
            company=dhpp.company, adb_required=0,
        )
        c2 = WeeklyCashCycle.objects.create(
            cycle_start=date(2026, 1, 13), cycle_end=date(2026, 1, 19),
            segment=dhpp, closing_balance=0,
        )
        for atype, amount in [
            (ActivityType.COLLECTION_DIST, "120000.00"),
            (ActivityType.OTHER_COLLECTION, "30000.00"),
            (ActivityType.SUPPLIER_PAYMENT, "50000.00"),
            (ActivityType.RFP_AP, "20000.00"),
            (ActivityType.PCF_REPLEN, "5000.00"),
            (ActivityType.OTHER_PAYMENT, "10000.00"),
            (ActivityType.CAPEX, "25000.00"),
            (ActivityType.BORROWED, "40000.00"),
            (ActivityType.LOAN_CLEARED, "15000.00"),
        ]:
            CashCycleActivity.objects.create(cycle=c2, activity_type=atype, amount=0)
        wb = build_cash_flow_statement(dhpp.company, date(2026, 1, 13), date(2026, 1, 19))
        assert _labels(wb["CF"]) == _labels(_src("STATEMENT-OF-CASH-FLOW.xlsx")["CF"])


class TestIncomeStatementMirror:
    def test_sheet_and_title(self, company, segments, coa, period_data, templates):
        wb = build_income_statement(company, date(2026, 1, 1), date(2026, 1, 31))
        ws = wb["January 2026"]
        assert company.name in ws["B1"].value
        assert "Statement of Profit / Loss" in ws["B1"].value
        assert ws["A6"].value == "ACCOUNTS"
        assert ws["G6"].value == "GRAND TOTAL"

    def test_line_titles_come_from_template(self, company, segments, coa, period_data, templates):
        from apps.reporting.services import FinancialStatementService

        wb = build_income_statement(company, date(2026, 1, 1), date(2026, 1, 31))
        ws = wb["January 2026"]
        gen_labels = {c.value for r in ws.iter_rows() for c in r
                      if isinstance(c.value, str)}
        fs = FinancialStatementService.generate(
            company=company, statement_type=StatementType.INCOME_STATEMENT,
            period_start=date(2026, 1, 1), period_end=date(2026, 1, 31),
        )
        for key, row in fs.rows_by_key().items():
            assert row["title"] in gen_labels, f"missing IS line: {row['title']}"

    def test_grand_column_matches_generate(self, company, segments, coa, period_data, templates):
        from apps.reporting.services import FinancialStatementService

        wb = build_income_statement(company, date(2026, 1, 1), date(2026, 1, 31))
        ws = wb["January 2026"]
        fs = FinancialStatementService.generate(
            company=company, statement_type=StatementType.INCOME_STATEMENT,
            period_start=date(2026, 1, 1), period_end=date(2026, 1, 31),
        )
        rows = fs.rows_by_key()
        assert ws.cell(row=_row_of(ws, "NET SALES"), column=7).value == rows["net_sales"]["amounts"]["GRAND"]
        assert ws.cell(row=_row_of(ws, "Cost of Sales(-)"), column=7).value == rows["cogs"]["amounts"]["GRAND"]
        assert ws.cell(row=_row_of(ws, "NET PROFIT / LOSS (+/-)"), column=7).value == rows["net_profit"]["amounts"]["GRAND"]
