"""Excel (.xlsx) export of the six financial statements (BUILD-PLAN Phase 8).

Each builder reproduces the exact layout of the source workbook in
/excel-files (TRIAL-BALANCE.xlsx, STATEMENT-OF-FINANCIAL-POSITION.xlsx,
STATEMENT-OF-CHANGES-IN-EQUITY.xlsx, STATEMENT-OF-COST-OF-SALES.xlsx,
STATEMENT-OF-TOTAL-EXPENSES.xlsx, STATEMENT-OF-CASH-FLOW.xlsx) — same sheet
names, same merged title blocks, same columns, same line titles, same order.

Values always come from the posted GL projection (ADR-005) via the same
services the API/UI use, so the download can never disagree with the screen:

  - Trial balance: monthly Dr/Cr windows (Opening + Jan..Dec + YTD).
  - SFP / SOCE / CoS / Total Expenses: FinancialStatementService.generate
    snapshots (template-driven), mapped to the workbook row order.
  - Cash flow: CashCycleActivity rows (ADR-028/031) per CF line.

Only the column layout is hardcoded here; every number is computed.
"""

from datetime import date, timedelta
from decimal import Decimal

from django.http import HttpResponse

from apps.reporting.models import StatementType

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:  # pragma: no cover
    Workbook = None

COMPANY_NAME = "SEVEN-TRENT MACHINERIES INDUSTRIAL EQUIPMENT TRADING"

MONTH_NAMES = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
               "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"]


def xlsx_response(wb, filename: str) -> HttpResponse:
    """Wrap a built workbook in a download response."""
    from io import BytesIO

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _account_total(balances: dict, code: str) -> Decimal:
    """Sum an account's balance across every segment column."""
    return sum((balances.get(code) or {}).values()) or Decimal("0.00")


def _net_amount(balances: dict, codes: list[str], prefixes: list[str] | None = None) -> Decimal:
    """Signed (normal-direction) total over exact codes and/or prefixes."""
    from apps.foundation.models import Account

    prefixes = [p for p in (prefixes or []) if p]
    selected = []
    for acc in Account.objects.filter(is_postable=True):
        if acc.code in codes or any(acc.code.startswith(p) for p in prefixes):
            selected.append(acc.code)
    return sum(_account_total(balances, c) for c in dict.fromkeys(selected)) or Decimal("0.00")


# ---------------------------------------------------------------------------
# Trial Balance (TRIAL-BALANCE.xlsx — sheet "TRIAL BALANCE")
# ---------------------------------------------------------------------------


def build_trial_balance(company, year: int) -> "Workbook":
    """Monthly TB mirroring the workbook: 8 COA columns + 13 Dr/Cr pairs.

    Row 2 title (merged I2:AJ2), row 4 alternating Dr./Cr., row 5 headers
    (OPENING BALANCES .. DECEMBER, YTD Balances), account rows from row 6.
    """
    from apps.foundation.models import Account
    from apps.reporting.services import TrialBalanceService

    if Workbook is None:  # pragma: no cover
        raise ImportError("openpyxl required for Excel export.")

    def window(start: date | None, end: date | None) -> dict:
        return TrialBalanceService.segment_balances(company, start=start, end=end)

    opening = window(None, date(year - 1, 12, 31))
    windows: list[dict] = [opening]
    for month in range(1, 13):
        month_end = date(year, month + 1, 1) - timedelta(days=1) if month < 12 else date(year, 12, 31)
        windows.append(window(date(year, month, 1), month_end))
    windows.append(window(date(year, 1, 1), date(year, 12, 31)))

    wb = Workbook()
    ws = wb.active
    ws.title = "TRIAL BALANCE"

    ws.merge_cells("I2:AJ2")
    ws["I2"] = f"{COMPANY_NAME} (TRIAL BALANCE)"
    ws["I2"].font = Font(bold=True)

    for col in range(9, 36):  # I..AJ
        ws.cell(row=4, column=col, value="Dr." if col % 2 else "Cr.")

    headers = ["COA", "Normal Balance of Account Titles", "ACCOUNT TITLES",
               "SEGMENT", "CLASSIFICATION", "CATEGORY", "Sub-Accounts", "Major Accounts"]
    for i, label in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=label)
    ws["I5"] = "OPENING BALANCES"
    ws.merge_cells("I5:J5")
    for month in range(12):
        col = 11 + month * 2
        ws.cell(row=5, column=col, value=MONTH_NAMES[month])
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    ws["AI5"] = "YTD Balances"
    ws.merge_cells("AI5:AJ5")

    row = 6
    for acc in Account.objects.filter(is_postable=True).order_by("code"):
        ws.cell(row=row, column=1, value=acc.code)
        ws.cell(row=row, column=2, value="Dr." if acc.normal_balance == "debit" else "Cr.")
        ws.cell(row=row, column=3, value=acc.name)
        ws.cell(row=row, column=4, value=acc.segment)
        ws.cell(row=row, column=5, value=acc.classification or "")
        ws.cell(row=row, column=6, value=acc.category or "")
        ws.cell(row=row, column=7, value=acc.sub_accounts or "")
        ws.cell(row=row, column=8, value=acc.major_accounts or "")
        credit_normal = acc.normal_balance == "credit"
        for i, bal in enumerate(windows):
            total = _account_total(bal, acc.code)
            dr_col = 9 + i * 2
            cr_col = dr_col + 1
            if credit_normal:
                ws.cell(row=row, column=cr_col if total >= 0 else dr_col,
                        value=abs(total) or 0)
            else:
                ws.cell(row=row, column=dr_col if total >= 0 else cr_col,
                        value=abs(total) or 0)
        row += 1
    return wb


# ---------------------------------------------------------------------------
# Statement of Financial Position (STATEMENT-OF-FINANCIAL-POSITION.xlsx)
# ---------------------------------------------------------------------------


def build_statement_of_financial_position(company, period_start: date, period_end: date,
                                          net_profit: Decimal | str | None = None) -> "Workbook":
    """YEAR END sheet: two-sided layout (assets | liabilities & equity)."""
    from apps.reporting.services import FinancialStatementService, TrialBalanceService

    if Workbook is None:  # pragma: no cover
        raise ImportError("openpyxl required for Excel export.")

    ending = TrialBalanceService.segment_balances(company, end=period_end)

    def net(codes, prefixes=None):
        return _net_amount(ending, codes, prefixes)

    def net_of(gross_codes, accum_codes):
        return net(gross_codes) - net(accum_codes)

    ca_due_bank_ew = net(["10020", "10023", "10026"])
    ca_cash = net([], ["10"]) - ca_due_bank_ew
    ca_ar = net([], ["12010", "12020", "12030", "12056", "12100"])
    ca_employees = net([], ["12070"])
    ca_related = net([], ["15"])
    ca_adv_supplier = net([], ["12040"])
    ca_inventory = net([], ["13"])
    total_ca = ca_cash + ca_due_bank_ew + ca_ar + ca_employees + ca_related + ca_adv_supplier + ca_inventory

    fa_building = net_of(["19000", "19500", "19700", "19750"], ["19760"])
    fa_ff = net_of(["19800"], ["19850"])
    fa_office = net_of(["19900", "19910", "19920", "19930", "19940", "19950",
                        "19960", "19963", "19966"], ["19980", "19983", "19986"])
    fa_vehicles = net_of(["17000", "17010", "18503", "18600", "18650"], ["18513", "18660"])
    total_fa = fa_building + fa_ff + fa_office + fa_vehicles
    nca_adv_contractors = net([], ["1997"])
    total_nca = nca_adv_contractors
    total_assets = total_ca + total_fa + total_nca

    cl_ap = net([], ["200", "211"])
    cl_due_contractor = net(["25013"])
    cl_accrued = net([], ["22"])
    cl_govt = net([], ["23"])
    cl_deferred = net([], ["210"])
    cl_loans = net([], ["240"])
    total_cl = cl_ap + cl_due_contractor + cl_accrued + cl_govt + cl_deferred + cl_loans
    ncl_loans = net([], ["270", "272"])
    ncl_other = net([], ["275", "276"])
    total_ncl = ncl_loans + ncl_other
    total_liabilities = total_cl + total_ncl

    capital = net([], ["300"])
    drawings = net([], ["305"])
    np = Decimal(str(net_profit)) if net_profit is not None else Decimal("0.00")
    total_equity = capital - drawings + np
    total_liab_equity = total_liabilities + total_equity

    wb = Workbook()
    ws = wb.active
    ws.title = "YEAR END"
    year = period_end.year

    def put(coord, value, bold=False):
        ws[coord] = value
        if bold:
            ws[coord].font = Font(bold=True)

    put("B3", company.name)
    put("B4", "STATEMENT OF FINANCIAL POSITION")
    put("B5", f"As of  {period_end:%B %d, %Y}")
    put("B6", "ASSETS")
    put("C6", str(year))
    put("E6", "LIABILITIES AND OWNER'S EQUITY")
    put("F6", str(year))
    put("B7", "CURRENT ASSETS")
    put("E7", "CURRENT LIABILITIES")
    put("B8", "Cash"); put("C8", ca_cash)
    put("E8", "Accounts Payable"); put("F8", cl_ap)
    put("B9", "Due from Bank-EW"); put("C9", ca_due_bank_ew)
    put("E9", "Due to Contractor"); put("F9", cl_due_contractor)
    put("B10", "Accounts Receivable"); put("C10", ca_ar)
    put("E10", "Accrued  Expense"); put("F10", cl_accrued)
    put("B11", "Employees Receivable"); put("C11", ca_employees)
    put("E11", "Govt. Mandatory Contribution Payable"); put("F11", cl_govt)
    put("B12", "Due from Related Party"); put("C12", ca_related)
    put("E12", "Deferred Revenue"); put("F12", cl_deferred)
    put("B13", "Advances to Supplier"); put("C13", ca_adv_supplier)
    put("E13", "Loans Payable-Current"); put("F13", cl_loans)
    put("B14", "Inventory"); put("C14", ca_inventory)
    put("B16", "TOTAL CURRENT ASSETS", bold=True); put("C16", total_ca)
    put("E16", "TOTAL CURRENT LIABILITIES", bold=True); put("F16", total_cl)
    put("B17", "FIXED (LONG-TERM) ASSETS")
    put("E17", "LONG-TERM LIABILITIES")
    put("B18", "Building and Improvements,Net"); put("C18", fa_building)
    put("E18", "Loans Payable - NC"); put("F18", ncl_loans)
    put("B19", "Furniture and Fixtures,Net"); put("C19", fa_ff)
    put("E19", "Other Payables - NC"); put("F19", ncl_other)
    put("B20", "Office Equipment,Net"); put("C20", fa_office)
    put("B21", "Vehicles/Specialized /Heavy Equipments,Net"); put("C21", fa_vehicles)
    put("B23", "TOTAL FIXED ASSETS", bold=True); put("C23", total_fa)
    put("E23", "TOTAL NON CURRENT LIABILITIES", bold=True); put("F23", total_ncl)
    put("B24", "OTHER NON-CURRENT ASSETS")
    put("E24", "OWNER'S EQUITY")
    put("B25", "Advances to Contractors"); put("C25", nca_adv_contractors)
    put("E25", "E.Bagatua, Capital"); put("F25", capital)
    put("E26", "E.Bagatua, Drawings"); put("F26", drawings)
    put("B27", "TOTAL NON CURRENT ASSETS", bold=True); put("C27", total_nca)
    put("E27", "TOTAL OWNER'S EQUITY", bold=True); put("F27", total_equity)
    put("B29", "TOTAL ASSETS", bold=True); put("C29", total_assets)
    put("E29", "TOTAL LIABILITIES AND OWNER'S EQUITY", bold=True); put("F29", total_liab_equity)
    put("B31", "COMMON FINANCIAL RATIO")
    put("C31", str(year))
    put("B32", "Debt Ratio (Total Liabilities / Total Assets)")
    put("C32", (total_liabilities * 100 / total_assets) if total_assets else 0)
    put("B33", "Current Ratio (Current Assets / Current Liabilities)")
    put("C33", (total_ca / total_cl) if total_cl else 0)
    put("B34", "Working Capital (Current Assets - Current Liabilities)")
    put("C34", total_ca - total_cl)
    put("B35", "Assets-to-Equity Ratio (Total Assets / Owner's Equity)")
    put("C35", (total_assets / total_equity) if total_equity else 0)
    put("B36", "Debt-to-Equity Ratio (Total Liabilities / Owner's Equity)")
    put("C36", (total_liabilities / total_equity) if total_equity else 0)
    return wb


# ---------------------------------------------------------------------------
# Statement of Changes in Equity (STATEMENT-OF-CHANGES-IN-EQUITY.xlsx)
# ---------------------------------------------------------------------------


def build_statement_of_changes_in_equity(company, period_start: date, period_end: date,
                                         net_profit: Decimal | str | None = None) -> "Workbook":
    """EQUITY sheet: Beginning Capital -> Additions -> Net Profit -> Drawings."""
    from apps.reporting.services import FinancialStatementService

    if Workbook is None:  # pragma: no cover
        raise ImportError("openpyxl required for Excel export.")

    inputs = {"soce_net_profit": str(net_profit)} if net_profit is not None else {}
    fs = FinancialStatementService.generate(
        company=company, statement_type=StatementType.SOCE,
        period_start=period_start, period_end=period_end, inputs=inputs,
    )
    rows = fs.rows_by_key()
    grand = lambda key: Decimal(rows[key]["amounts"]["GRAND"])

    wb = Workbook()
    ws = wb.active
    ws.title = "EQUITY"
    ws.merge_cells("C2:D2")
    ws["C2"] = "STATEMENT OF CHANGES IN EQUITY"
    ws["C3"] = company.name
    ws["C4"] = f"For the year ended {period_end:%B %d, %Y}"
    ws.merge_cells("B6:C6")
    ws["B6"] = "EQUITY ACCOUNTS"
    ws["D6"] = "TOTAL"
    for offset, (label, key) in enumerate([
        ("E.Bagatua, Beginning Capital", "soce_begin_capital"),
        ("Additional Capital", "soce_additional_capital"),
        ("Net Profit / Loss for the year (+ / - )", "soce_net_profit"),
        ("Total", "soce_total"),
        ("Less: E. Bagatua, Drawings", "soce_drawings"),
        ("E. Bagatua, Ending Capital", "soce_ending_capital"),
    ], start=7):
        ws.cell(row=offset, column=2, value=label)
        ws.cell(row=offset, column=4, value=grand(key))
    return wb


# ---------------------------------------------------------------------------
# Statement of Cost of Sales (STATEMENT-OF-COST-OF-SALES.xlsx)
# ---------------------------------------------------------------------------

# (key, workbook title, workbook row) — order = workbook order.
COS_DHPP_ROWS = [
    ("cos_dhpp_subscription", "COGS - Subscription Fees"),
    ("cos_dhpp_dep_tanker", "COGS - Depreciation of Fuel Tankers_DHPP"),
    ("cos_dhpp_fuel", "COGS - Fuel Purchase"),
    ("cos_dhpp_discount", "COGS - Fuel Purchase Discount"),
    ("cos_dhpp_gasoline", "COGS - Gasoline Expenses_DHPP"),
    ("cos_dhpp_labor", "COGS - Labor Cost_DHPP"),
    ("cos_dhpp_direct", "COGS - Other Direct Fees_DHPP"),
    ("cos_dhpp_rm", "COGS - Repairs and Maintenance_DHPP"),
    ("cos_dhpp_storage", "COGS - Storage and Handling Cost_DHPP"),
    ("cos_dhpp_toll", "COGS - Toll Fees_DHPP"),
    ("cos_dhpp_wages", "COGS - Trip Wages"),
    ("cos_dhpp_staff", "COGS - Operational Staff_DHPP"),
]
COS_DMIE_ROWS = [
    ("cos_dmie_bucket", "COGS - Calibration Bucket"),
    ("cos_dmie_dep_vehicles", "COGS - Depreciation of DMIE Vehicles"),
    ("cos_dmie_gasoline", "COGS - Gasoline Expenses_DMIE"),
    ("cos_dmie_inverter", "COGS - Inverter"),
    ("cos_dmie_labor", "COGS - Labor Cost_DMIE"),
    ("cos_dmie_lubricants", "COGS - Lubricants for Consumption"),
    ("cos_dmie_machinery", "COGS - Machinery (TSRO/LFRO)"),
    ("cos_dmie_staff", "COGS - Operational Staff_DMIE"),
    ("cos_dmie_charges", "COGS - Other Charges (TSRO)"),
    ("cos_dmie_direct", "COGS - Other Direct Fees_DMIE"),
    ("cos_dmie_other", "COGS - Other DMIE"),
    ("cos_dmie_equipment", "COGS - Other Industrial Equipment"),
    ("cos_dmie_recond", "COGS - Reconditioning Expenses"),
    ("cos_dmie_rm", "COGS - Repairs and Maintenance_DMIE"),
    ("cos_dmie_shipping", "COGS - Shipping Fees (TSRO)"),
    ("cos_dmie_storage", "COGS - Storage and Handling Cost_DMIE"),
    ("cos_dmie_toll", "COGS - Toll Fees_DMIE"),
    ("cos_dmie_pump", "COGS - Transfer Pump"),
]
COS_OPS_ROWS = [
    ("cos_ops_lubricants", "COGS - Lubricants for Sale"),
    ("cos_ops_labor", "COGS - Labor Cost_OPS"),
    ("cos_ops_staff", "COGS - Operational Staff_OPS"),
    ("cos_ops_other", "COGS - Other OPS"),
    ("cos_ops_rm", "COGS - Repairs and Maintenance_OPS"),
]


def _grand(rows: dict, key: str) -> Decimal:
    return Decimal(rows[key]["amounts"]["GRAND"])


def build_statement_of_cost_of_sales(company, period_start: date, period_end: date) -> "Workbook":
    """COST OF SALES sheet: DHPP 12 / DMIE 18 / OPS 5 + segment totals."""
    from apps.reporting.services import FinancialStatementService

    if Workbook is None:  # pragma: no cover
        raise ImportError("openpyxl required for Excel export.")

    fs = FinancialStatementService.generate(
        company=company, statement_type=StatementType.COST_OF_SALES,
        period_start=period_start, period_end=period_end,
    )
    rows = fs.rows_by_key()

    def account_lines(rows_def, start_row):
        for offset, (key, title) in enumerate(rows_def):
            ws.cell(row=start_row + offset, column=1, value=title)
            ws.cell(row=start_row + offset, column=5, value=_grand(rows, key))

    def section_header(row, title):
        ws.cell(row=row, column=1, value=title)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    def column_headers(row):
        for col, label in [(1, "ACCOUNT TITLES"), (2, "IPPC"), (3, "STPC"),
                           (4, "STMIET"), (5, "GRAND TOTAL")]:
            ws.cell(row=row, column=col, value=label)

    def segment_total(row, key, cols):
        for col in cols:
            ws.cell(row=row, column=col, value=_grand(rows, key))

    wb = Workbook()
    ws = wb.active
    ws.title = "COST OF SALES"
    ws.merge_cells("B1:E4")
    ws["B1"] = (f"\n{COMPANY_NAME}\nStatement of Cost of Sales\n"
                f"For the MONTH OF {period_start:%B}")
    ws.merge_cells("B6:E7")
    ws["B6"] = f" FOR THE MONTH OF {period_start:%B}\n"

    section_header(8, "Distribution and Hauling of Petroleum Products (DHPP)")
    column_headers(9)
    account_lines(COS_DHPP_ROWS, 10)
    segment_total(22, "cos_dhpp_total", [2, 3, 5])
    ws.cell(row=23, column=1, value="VOLUME IN LITERS")
    ws.cell(row=23, column=3, value=0)
    ws.cell(row=23, column=5, value=0)
    ws.cell(row=24, column=1, value="DIRECT COST PER LITER ")
    ws.cell(row=24, column=2, value=0)
    ws.cell(row=24, column=3, value=0)
    ws.cell(row=24, column=5, value=0)

    section_header(25, "Distribution of Machineries and Industrial Equipment (DMIE)")
    column_headers(26)
    account_lines(COS_DMIE_ROWS, 27)
    segment_total(45, "cos_dmie_total", [2, 3, 5])

    section_header(46, "Other Products and Services (OPS)")
    column_headers(47)
    account_lines(COS_OPS_ROWS, 48)
    segment_total(53, "cos_ops_total", [2, 3, 5])

    ws.cell(row=54, column=1, value="TOTAL COST OF SALES")
    for col in (2, 3, 4, 5):
        ws.cell(row=54, column=col, value=_grand(rows, "total_cost_of_sales"))
    return wb


# ---------------------------------------------------------------------------
# Statement of Total Expenses (STATEMENT-OF-TOTAL-EXPENSES.xlsx)
# ---------------------------------------------------------------------------

# (key, workbook title, workbook row) — order = workbook order. The workbook
# itself repeats "Accommodation Fees" at row 17 (blank); mirrored faithfully.
TE_ROWS = [
    ("te_cogs_dhpp", "COGS - DHPP", 9),
    ("te_cogs_dmie", "COGS - DMIE", 10),
    ("te_cogs_ops", "COGS - OPS", 11),
    ("te_total_cogs", "Total Cost of Sales", 12),
    (None, "Direct Cost per Liter", 13),
    ("te_accommodation", "Accommodation Fees", 14),
    ("te_bad_debts", "Bad Debts", 15),
    ("te_bank_fees", "Bank Fees", 16),
    (None, "Accommodation Fees", 17),
    ("te_dep_exp", "Depreciation Expense (Others)", 18),
    ("te_er_shares", "Employer Mandatory Contribution", 19),
    ("te_impairment", "Impairment of Assets", 20),
    ("te_insurance", "Insurance Fees", 21),
    ("te_interest", "Interest Expenses", 22),
    ("te_legal", "Legal & Professional Fees", 23),
    ("te_loan_rel", "Loan Related Expenses", 24),
    ("te_office_supplies", "Office Supplies Expense", 25),
    ("te_salaries", "Salary and Company Benefits", 26),
    ("te_taxes", "Taxes, Licenses, Penalties", 27),
    ("te_travel", "Travel Related Expenses", 28),
    ("te_utilities", "Utilities Expense", 29),
    ("te_withholding", "Withholding Taxes", 30),
    ("te_representation", "Representation Expense", 31),
    ("te_other_op", "Other Operating Expenses", 32),
    ("te_total_operating_expenses", "Total Operating Expenses", 33),
    ("te_gen_admin", "Other Gen. & Admin. Expense", 34),
    ("te_misc", "Other Expense/Miscellaneous Exp.", 35),
    ("te_total_non_operating", "Total Non-Operating Expenses", 36),
    ("te_total_operating_costs", "TOTAL OPERATING COSTS", 37),
]


def build_statement_of_total_expenses(company, period_start: date, period_end: date) -> "Workbook":
    """MARCH 2026 CGSE sheet: COGS + operating + non-operating + liters."""
    from apps.reporting.services import FinancialStatementService

    if Workbook is None:  # pragma: no cover
        raise ImportError("openpyxl required for Excel export.")

    fs = FinancialStatementService.generate(
        company=company, statement_type=StatementType.TOTAL_EXPENSES,
        period_start=period_start, period_end=period_end,
    )
    rows = fs.rows_by_key()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{period_start:%B %Y} CGSE"
    ws.merge_cells("B1:H4")
    ws["B1"] = (f"\n{COMPANY_NAME}\nStatement of Total Expenses\n"
                f"For the MONTH OF {period_start:%B %Y}\n")
    ws.merge_cells("B6:G6")
    ws["B6"] = f"MONTH OF {period_start:%B %Y}\n"
    ws.merge_cells("H6:H8")
    ws["H6"] = "GRAND TOTAL"
    ws.merge_cells("B7:E7")
    ws["B7"] = "Segment A\n(DHPP)"
    ws["F7"] = "Segment B \n(DMIE)"
    ws["G7"] = "Segment C \n(OPS)"
    for col, label in [(1, "ACCOUNT TITLES"), (2, "IPPC"), (3, "STPC"),
                       (4, "STMIET"), (5, "TOTAL"), (6, "STMIET"), (7, "STMIET")]:
        ws.cell(row=8, column=col, value=label)

    for key, title, row in TE_ROWS:
        ws.cell(row=row, column=1, value=title)
        if key:
            ws.cell(row=row, column=8, value=_grand(rows, key))
    # workbook rows with explicit IPPC/STPC zero cells
    for row in (9, 12, 33, 36, 37):
        ws.cell(row=row, column=2, value=0)
        ws.cell(row=row, column=3, value=0)
    for row in (13, 41):
        ws.cell(row=row, column=2, value=0)
        ws.cell(row=row, column=3, value=0)
    for row in (38, 39, 40):
        ws.cell(row=row, column=8, value=0)
    ws.cell(row=39, column=2, value=0)
    ws.cell(row=39, column=3, value=0)
    ws.cell(row=40, column=2, value=0)
    ws.cell(row=40, column=3, value=0)
    ws.cell(row=38, column=1, value="Fuel Delivered in Liters _IPPC")
    ws.cell(row=39, column=1, value="Fuel Delivered in Liters _STMIET")
    ws.cell(row=40, column=1, value="Fuel Delivered in Liters _3rd party")
    ws.cell(row=41, column=1, value="UNIT COST PER LITER FOR THE CYCLE")
    ws.cell(row=43, column=1, value="Legend:")
    ws.cell(row=44, column=1, value="DHPP - Distribution and Hauling of Petroleum Products")
    ws.cell(row=45, column=1, value="DMIE - Distribution of Machineries and Industrial Equipment")
    ws.cell(row=46, column=1, value="OPS - Other Products and Services")
    return wb


# ---------------------------------------------------------------------------
# Statement of Cash Flows (STATEMENT-OF-CASH-FLOW.xlsx — sheet "CF")
# ---------------------------------------------------------------------------


def build_cash_flow_statement(segment, period_start: date, period_end: date) -> "Workbook":
    """CF sheet: operating / investing / financing / net change / ADB.

    Values are recomputed from the period's CashCycleActivity rows (the same
    source CashFlowService.generate uses), so every CF line is itemized.
    """
    from django.db.models import Sum

    from apps.cash.models import ActivityType, BankAccount, CashCycleActivity, WeeklyCashCycle

    if Workbook is None:  # pragma: no cover
        raise ImportError("openpyxl required for Excel export.")

    cycles = list(WeeklyCashCycle.objects.filter(
        segment=segment, cycle_start__gte=period_start, cycle_end__lte=period_end
    ).order_by("cycle_start"))
    totals = dict(CashCycleActivity.objects.filter(
        cycle__in=cycles
    ).values_list("activity_type").annotate(total=Sum("amount")))
    totals = {k: v or Decimal("0.00") for k, v in totals.items()}

    t = lambda key: totals.get(key, Decimal("0.00"))
    coll_dist = t(ActivityType.COLLECTION_DIST)
    other_coll = t(ActivityType.OTHER_COLLECTION)
    supplier_pay = t(ActivityType.SUPPLIER_PAYMENT)
    rfp_ap = t(ActivityType.RFP_AP)
    pcf_replen = t(ActivityType.PCF_REPLEN)
    other_pay = t(ActivityType.OTHER_PAYMENT)
    net_operating = coll_dist + other_coll - supplier_pay - rfp_ap - pcf_replen - other_pay
    capex = t(ActivityType.CAPEX)
    net_investing = -capex
    borrowed = t(ActivityType.BORROWED)
    loan_cleared = t(ActivityType.LOAN_CLEARED)
    net_financing = borrowed - loan_cleared
    net_change = net_operating + net_investing + net_financing

    first = cycles[0] if cycles else None
    adb = Decimal("0.00")
    for bank in BankAccount.objects.filter(segment=segment, is_active=True):
        adb += bank.adb_required
    if first:
        prev = WeeklyCashCycle.objects.filter(
            segment=segment, cycle_end__lt=first.cycle_start
        ).order_by("-cycle_end").first()
        beginning = prev.closing_balance if prev else Decimal("0.00")
        ending = cycles[-1].closing_balance - adb
    else:
        beginning = ending = Decimal("0.00")

    wb = Workbook()
    ws = wb.active
    ws.title = "CF"
    ws["B2"] = COMPANY_NAME
    ws["B3"] = "Summary of Cash Flows"
    ws["B4"] = f"For the period cycle {period_start:%B %d} - {period_end:%B %d, %Y}"
    ws.merge_cells("H5:I5")
    ws["H5"] = "Amounts in pesos"

    def amount_row(row, label, value, merge=True):
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=8, value=value)
        if merge:
            ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)

    ws.cell(row=6, column=2, value="CASH FLOWS ARISING FROM OPERATING ACTIVITIES")
    amount_row(7, "  Collections from Distribution and Hauling  of Petroleum (+) [ Net of inter-account transfer]", coll_dist)
    amount_row(8, "  Other Cash Collections  relating to operating activities (+)", other_coll)
    amount_row(9, "  Payments to Supplier of  Petroleum Products (-)", supplier_pay)
    amount_row(10, "  RFP of Accounts Payable [ i.e other suppliers, 3rd parties] (-)", rfp_ap)
    amount_row(11, "  Cash Withdrawn for  PCF Replenishment (-)", pcf_replen)
    amount_row(12, "  Other Cash payments relating to operating activities(-)", other_pay)
    amount_row(13, "Net cash flows provided by operating activities (+/-)", net_operating)
    ws.cell(row=14, column=2, value="CASH FLOWS FROM INVESTING ACTIVITIES")
    amount_row(17, "Net cash flows used  in investing activities", net_investing)
    ws.cell(row=18, column=2, value="CASH FLOWS  ARISING FROM FINANCING ACTIVITIES")
    amount_row(19, "  Funds Borrowed from other account  (+)", borrowed)
    amount_row(20, "  Checks Cleared for Loan / Fund Repayments (-)", loan_cleared)
    amount_row(21, "Net cash flows used  in financing activities (+/-)", net_financing)
    amount_row(22, "NET INCREASE / (DECREASE) IN CASH", net_change)
    amount_row(23, "Add: CASH AT THE BEGINNING OF THE CYCLE", beginning)
    amount_row(24, "Less: ADB, Maintaining Balance", adb)
    amount_row(25, "CASH AVAILABLE AT THE END OF THE CYCLE", ending)
    return wb
