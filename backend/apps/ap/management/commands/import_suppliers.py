"""Bulk-import the supplier master (BUILD-PLAN Phase 10 master-data migration).

Usage:
    py manage.py import_suppliers --file excel-files/SUPPLIERS.xlsx
    py manage.py import_suppliers --file excel-files/SUPPLIERS.csv

Two supported layouts (columns are located by header name, not position):

  1. Legacy master:
       CODE | NAME | TYPE | TIN | ADDRESS | CONTACT | DEFAULT SEGMENT | LAST AP

       - TYPE            : depot | equipment | service | govt | other (defaults other)
       - DEFAULT SEGMENT : "DHPP", "DMIE", "OPS" (optional)
       - LAST AP         : previous RFP number for per-vendor gap tracking (optional)

  2. Finance head's LIST-OF-SUPPLIERS (Sept 01 2026) master:
       No sort code column; column A is a display sequence number and is ignored.
       Header row is detected by the "BUSINESS NAME" header. A supplier may span
       two physical rows (primary + secondary contact person); these are merged
       into one Supplier record, so the secondary contact/position/phones are
       preserved rather than creating duplicates.

       BUSINESS NAME | OWNER/REP/PRESIDENT | EMAIL ADDRESS | TIN | BUSINESS ADDRESS
         | CONTACT PERSON | POSITION | CONTACT NUMBERS | ATTACHMENT NEEDED

  A stable, unique `code` is derived from the business name when the source has
  no explicit supplier/vendor code (per-vendor LAST-AP numbering needs a key).

Idempotent: a CODE that exists is updated in place; re-runs are safe.
"""

import csv
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.ap.models import Supplier, SupplierType
from apps.foundation.models import Company, Segment

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

HEADER_MAP = {
    "code": ("CODE", "SUPPLIER CODE", "VENDOR CODE", "ACCT #", "ACC #"),
    "name": (
        "NAME", "SUPPLIER NAME", "VENDOR", "PAYEE",
        "BUSINESS NAME", "BUSINESS", "TRADE NAME",
    ),
    "sup_type": ("TYPE", "SUPPLIER TYPE", "CATEGORY", "KIND"),
    "tin": ("TIN",),
    "address": ("ADDRESS", "BUSINESS ADDRESS"),
    "contact": ("CONTACT", "CONTACT NO", "CONTACT NOS", "CONTACT NUMBERS", "PHONE", "MOBILE"),
    "owner_name": ("OWNER/REP/PRESIDENT", "OWNER", "REPRESENTATIVE",
                   "OWNER/REPRESENTATIVE/PRESIDENT", "OWNER/REPRESENTATIVE/ PRESIDENT", "PRESIDENT"),
    "email": ("EMAIL", "EMAIL ADDRESS"),
    "contact_person": ("CONTACT PERSON",),
    "position": ("POSITION", "DESIGNATION"),
    "attachments_required": ("ATTACHMENT NEEDED",
                             "ATTACHMENT NEEDED ( BIR-COR , DTI/SEC , BUSINESS PERMIT )",
                             "BIR-COR", "DTI/SEC", "PERMIT"),
    "default_segment": ("DEFAULT SEGMENT", "SEGMENT", "SECTION"),
    "last_ap": ("LAST AP", "PREVIOUS AP", "LAST AP NUMBER"),
}

TYPE_ALIASES = {
    "DEPOT": SupplierType.DEPOT, "FUEL": SupplierType.DEPOT, "DIESEL": SupplierType.DEPOT,
    "EQUIPMENT": SupplierType.EQUIPMENT, "MACHINERY": SupplierType.EQUIPMENT,
    "SERVICE": SupplierType.SERVICE, "SERVICES": SupplierType.SERVICE,
    "GOVT": SupplierType.GOVT, "GOVERNMENT": SupplierType.GOVT,
    "OTHER": SupplierType.OTHER,
}

SEGMENT_ALIASES = {
    "DHPP": "DHPP", "DMIE": "DMIE", "OPS": "OPS", "": None,
}


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Bulk-import the supplier master from CSV or XLSX (idempotent)."

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file", default=None, help="CSV/XLSX path")
        parser.add_argument("--company", dest="company", default="STMIET")

    def _load_rows(self, file_path):
        if file_path.suffix.lower() == ".csv":
            with open(file_path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.reader(fh)
                header = [h.strip().upper() for h in next(reader, [])]
                for i, row in enumerate(reader, start=2):
                    if not any(row):
                        continue
                    yield self._map_row(row, header), f"CSV row {i}"
        else:
            if openpyxl is None:
                raise CommandError("openpyxl required for xlsx input; pip install openpyxl")
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header = None
                header_idx = 0
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if not row or not any(row):
                        continue
                    upper = [str(v or "").strip().upper() for v in row]
                    if header is None and any("NAME" in h or "CODE" in h for h in upper):
                        header = upper
                        header_idx = i
                        break
                if header is None:
                    self.stdout.write(self.style.WARNING(
                        f"skip sheet {sheet_name}: no recognizable header row"
                    ))
                    continue
                for i, row in enumerate(ws.iter_rows(min_row=header_idx + 1, values_only=True), start=header_idx + 1):
                    if not row or not any(row):
                        continue
                    yield self._map_row(row, header), f"{sheet_name} row {i}"

    def _map_row(self, row, header):
        idx = {}
        for field, names in HEADER_MAP.items():
            for n in names:
                if n in header:
                    idx[field] = header.index(n)
                    break
        out = {}
        for field, pos in idx.items():
            if pos < len(row):
                out[field] = _clean(row[pos])
        return out

    def _resolve_type(self, raw):
        if not raw:
            return SupplierType.OTHER
        return TYPE_ALIASES.get(raw.upper().strip(), SupplierType.OTHER)

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = Path(options["file"]) if options["file"] else None
        if file_path is None or not file_path.exists():
            repo_root = Path(__file__).resolve().parents[5]
            for cand in (
                repo_root / "excel-files" / "SUPPLIERS.xlsx",
                repo_root / "excel-files" / "SUPPLIERS.csv",
                Path.cwd() / "excel-files" / "SUPPLIERS.xlsx",
            ):
                if cand.exists():
                    file_path = cand
                    break
        if file_path is None or not file_path.exists():
            raise CommandError(
                "Supplier master not found. Pass --file, or place SUPPLIERS.csv"
                "/.xlsx under /excel-files."
            )

        company = Company.objects.filter(code=options["company"]).first()
        if company is None:
            raise CommandError(f"Company '{options['company']}' not found. Run import_coa first.")

        created = updated = skipped = 0
        for record, label in self._group_suppliers(self._load_rows(file_path)):
            code = record.get("code", "")
            name = record.get("name", "")
            if not code and not name:
                skipped += 1
                continue
            try:
                if not code:
                    base = re.sub(r"[^A-Z0-9]+", "-", name.upper())[:16].strip("-")
                    code = f"{base}-{derived_seq()}"
                seg_code = SEGMENT_ALIASES.get(record.get("default_segment", "").upper().strip())
                default_segment = None
                if seg_code:
                    default_segment = Segment.objects.filter(code=seg_code).first()
                email = self._clean_email(record.get("email", ""))
                _, was_created = Supplier.objects.update_or_create(
                    code=code,
                    defaults={
                        "name": name or code,
                        "supplier_type": self._resolve_type(record.get("sup_type", "")),
                        "tin": record.get("tin", ""),
                        "address": record.get("address", ""),
                        "contact_no": record.get("contact", ""),
                        "owner_name": record.get("owner_name", ""),
                        "email": email,
                        "contact_person": record.get("contact_person", ""),
                        "position": record.get("position", ""),
                        "attachments_required": self._resolve_attachments(record.get("attachments_required", "")),
                        "default_segment": default_segment,
                        "last_ap": record.get("last_ap", ""),
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except CommandError as exc:
                self.stdout.write(self.style.ERROR(f"  {label}: {exc}"))
                skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f"import_suppliers: {created} created, {updated} updated, {skipped} skipped."
        ))

    @staticmethod
    def _resolve_attachments(raw):
        cleaned = _clean(raw).upper()
        if cleaned in ("", "NO", "NONE", "NOT REQUIRED"):
            return False
        return True

    @staticmethod
    def _clean_email(raw):
        cleaned = _clean(raw)
        if not cleaned or cleaned.count("@") != 1 or " " in cleaned:
            return ""
        return cleaned

    def _group_suppliers(self, rows):
        """Fold multi-row suppliers (primary + secondary contact rows) into one.

        A continuation row carries contact/position/phones but no business name
        (the source repeats the supplier on two physical rows). We merge those
        into the current record instead of emitting a stray empty-name row.
        """
        current = None
        current_label = None
        for row, label in rows:
            name = row.get("name", "")
            if name:
                if current is not None:
                    yield current, current_label
                current = dict(row)
                current_label = label
            elif current is not None:
                for field in ("contact_person", "position", "contact"):
                    extra = row.get(field, "")
                    if extra and extra not in str(current.get(field, "")):
                        current[field] = f"{current.get(field, '')} | {extra}".strip(" |")
        if current is not None:
            yield current, current_label


def derived_seq():
    n = getattr(derived_seq, "_n", 0) + 1
    derived_seq._n = n
    return f"{n:03d}"
