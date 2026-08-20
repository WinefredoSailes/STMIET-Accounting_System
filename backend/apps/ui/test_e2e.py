"""End-to-end workflow test: the full January-2026 happy path through the UI.

Walks the same flows the demo seed (`manage.py seed_demo`) creates but against
an isolated test DB, through the real HTTP views and bounded-context services:

    customer -> AR invoice -> collection JE -> RFP approval chain
    (incl. CNR escalation above P100k) -> CONSO batch post -> CV lifecycle
    -> inter-account transfer -> advance liquidation -> weekly cycles ->
    COLLECTIBLES + Cash Flow statement -> render the six register screens.

Any step failing here means the wiring between the UI and the services broke.
"""

from datetime import date, timedelta
from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model

User = get_user_model()


@pytest.fixture
def roles(db):
    """One active user per approval position: staff, head (Alywin), coo."""
    from apps.foundation.models import UserProfile

    out = {}
    for username in ("staff", "head", "coo"):
        out[username] = User.objects.create_user(username=username, password="x")
        UserProfile.objects.create(user=out[username], approval_role=username)
    return out


def create_invoice(segment, customer, txn_date, total):
    from apps.ar.models import ARInvoice, ARInvoiceLine

    inv = ARInvoice.objects.create(
        invoice_no=f"SI-E2E-{txn_date:%Y%m%d}",
        customer=customer,
        transaction_date=txn_date,
        segment=segment,
        total=Decimal(total),
        status="open",
    )
    ARInvoiceLine.objects.create(
        invoice=inv, line_no=1, product_code="DIESEL",
        description="Fuel delivery", quantity=1,
        unit_price=Decimal(total), amount=Decimal(total),
    )
    return inv


@pytest.fixture
def supplier(db, segment):
    from apps.ap.models import Supplier

    return Supplier.objects.create(
        code="E2E-SUP", name="E2E Fuel Depot", default_segment=segment
    )


@pytest.fixture
def doc_sequences(db, company):
    """ADR-032 numbering patterns, as configured in the dev database."""
    from apps.sequences.models import DocumentSequence

    for form_code, pattern in [
        ("AR", "AR-{YYYY}-{SEQ:05d}"),
        ("RFP", "A{SEQ:04d}"),
        ("CV", "CV-{YYYY}-{SEQ:04d}"),
        ("CONSO", "CONSO-{YYYY}-{SEQ:02d}"),
    ]:
        DocumentSequence.objects.create(
            company=company, form_code=form_code, year=2026, pattern=pattern
        )


@pytest.mark.django_db
class TestEndToEndWorkflow:

    def test_january_happy_path_renders_every_screen(
        self, client, company, segment, accounts, fiscal_period, user, roles, supplier, doc_sequences
    ):
        from apps.ar.models import ARInvoice, AcknowledgmentReceipt, Customer
        from apps.cash.models import (
            BankAccount, CashFlowStatement, CollectiblesWorksheet,
            InterAccountTransfer, WeeklyCashCycle,
        )
        from apps.cash.services import CashCycleService, CashFlowService, CollectiblesService
        from apps.posting.models import JournalEntry, PostingStatus

        staff = roles["staff"]
        client.force_login(staff)

        # 1. Customer master --------------------------------------------------
        resp = client.post(
            "/ar/customers/new/",
            {"code": "E2E-001", "name": "E2E Customer", "group": "fuel",
             "segment": segment.id, "pricing_tier": "regular"},
        )
        assert resp.status_code == 302
        customer = Customer.objects.get(code="E2E-001")

        # 2. AR invoice (no UI screen yet — created via the model) ------------
        inv = create_invoice(segment, customer, date(2026, 1, 6), "120000.00")

        # 3. Collection posted through the UI form -----------------------------
        resp = client.post(
            "/ar/receipts/new/",
            {"customer": customer.id, "cash_account": accounts["10010"].id,
             "transaction_date": "2026-01-07", "amount": "50000.00",
             "payment_method": "cash"},
        )
        assert resp.status_code == 302
        receipt = AcknowledgmentReceipt.objects.get(customer=customer)
        assert receipt.journal_entry.status == PostingStatus.POSTED
        assert (
            JournalEntry.objects.filter(status=PostingStatus.POSTED, source_doc_type="AR").count() == 1
        )

        # 4. RFP created + approved via the UI buttons -------------------------
        # 4a. Invalid submissions must NOT 500: blank line amount, blank date,
        # malformed amount, unbalanced Dr/Cr — the form re-renders with an error.
        from apps.ap.models import RFPDocument

        before = RFPDocument.objects.count()
        for bad in (
            {"line_amount": ["", "150000.00"], "line_side": ["dr", "cr"], "rfp_date": "2026-01-07"},
            {"line_amount": ["150000.00", "150000.00"], "line_side": ["dr", "cr"], "rfp_date": ""},
            {"line_amount": ["150000.00x", "150000.00"], "line_side": ["dr", "cr"], "rfp_date": "2026-01-07"},
            {"line_amount": ["150000.00", "50000.00"], "line_side": ["dr", "cr"], "rfp_date": "2026-01-07"},
        ):
            resp = client.post(
                "/ap/rfps/new/",
                {"payee": supplier.id, "segment": segment.id, **bad,
                 "purpose": "purchase",
                 "line_segment": [segment.id, segment.id],
                 "line_account": ["61100", "20000"],
                 "line_description": ["Fuel", "AP - E2E Fuel Depot"],
                 },
            )
            assert resp.status_code == 200, f"bad RFP posted {bad!r} -> {resp.status_code}"
        assert RFPDocument.objects.count() == before  # nothing persisted

        # 4b. Comma thousands are tolerated: money() strips separators.
        resp = client.post(
            "/ap/rfps/new/",
            {"payee": supplier.id, "segment": segment.id, "rfp_date": "2026-01-07",
             "purpose": "purchase",
             "line_segment": [segment.id, segment.id],
             "line_account": ["61100", "20000"],
             "line_amount": ["150,000.00", "150000.00"],
             "line_side": ["dr", "cr"],
             "line_description": ["Fuel", "AP - E2E Fuel Depot"],
             },
        )
        assert resp.status_code == 302
        comma_rfp = RFPDocument.objects.get(payee=supplier, amount=Decimal("150000.00"))
        assert comma_rfp.particulars == "Fuel"  # mirrors the first line description

        # 4c. Happy path: valid RFP created + approved step by step.
        resp = client.post(
            "/ap/rfps/new/",
            {"payee": supplier.id, "segment": segment.id, "rfp_date": "2026-01-07",
             "purpose": "purchase",
             "line_segment": [segment.id, segment.id],
             "line_account": ["61100", "20000"],
             "line_amount": ["150000.00", "150000.00"],
             "line_side": ["dr", "cr"],
             "line_description": ["E2E bulk fuel", "AP - E2E Fuel Depot"],
             },
        )
        assert resp.status_code == 302
        rfp = RFPDocument.objects.get(payee=supplier, particulars="E2E bulk fuel")
        assert rfp.status == "prepared"
        # 4d. Routing guard: the preparer (staff) has no step on a prepared
        # RFP and cannot approve it — the redirect keeps it in "prepared".
        client.force_login(roles["staff"])
        client.post(f"/ap/rfps/{rfp.id}/approve/")
        rfp.refresh_from_db()
        assert rfp.status == "prepared"
        # head (Alywin) checks then approves acctg + fin — two clicks,
        # two statuses, same person (ADR-036 relaxed same-user rule)
        client.force_login(roles["head"])
        client.post(f"/ap/rfps/{rfp.id}/approve/")
        client.post(f"/ap/rfps/{rfp.id}/approve/")
        client.post(f"/ap/rfps/{rfp.id}/approve/")
        rfp.refresh_from_db()
        assert rfp.status == "fin_approved"
        # CNR escalation (above P100k) — only the COO is a fresh hand
        client.force_login(roles["coo"])
        client.post(f"/ap/rfps/{rfp.id}/approve-cnr/")
        rfp.refresh_from_db()
        assert rfp.status == "cnr_approved"

        # 5. CONSO batch: open -> add RFP -> post -----------------------------
        client.force_login(roles["head"])
        resp = client.post("/ap/conso/new/", {"conso_date": "2026-01-16"})
        assert resp.status_code == 302
        from apps.ap.models import CONSOBatch

        batch = CONSOBatch.objects.get()
        client.post(f"/ap/conso/{batch.id}/add-rfp/", {"rfp": rfp.id})
        resp = client.post(f"/ap/conso/{batch.id}/post/")
        assert resp.status_code == 302
        batch.refresh_from_db()
        rfp.refresh_from_db()
        assert batch.status == "posted"
        assert rfp.status == "posted"
        assert rfp.journal_entry.status == PostingStatus.POSTED

        # 6. CV lifecycle through the UI --------------------------------------
        client.force_login(roles["staff"])
        resp = client.post(
            "/ap/cv/new/",
            {"rfp": rfp.id, "bank_account": accounts["10110"].id,
             "cv_date": "2026-01-20", "gross_amount": "60000.00",
             "withheld_tax": "6000.00", "check_no": "CHK-E2E"},
        )
        assert resp.status_code == 302
        from apps.ap.models import CheckVoucher

        cv = CheckVoucher.objects.get()
        assert cv.journal_entry.status == PostingStatus.POSTED
        client.force_login(roles["coo"])
        client.post(f"/ap/cv/{cv.id}/sign/")
        client.force_login(roles["head"])
        client.post(f"/ap/cv/{cv.id}/release/")
        client.post(f"/ap/cv/{cv.id}/clear/")
        cv.refresh_from_db()
        assert cv.status == "cleared"

        # 7. Inter-account transfer -------------------------------------------
        bank_from = BankAccount.objects.create(
            code="E2E-BDO", name="BDO E2E", account_type="checking",
            bank_name="BDO", bank_code="BDO", gl_account=accounts["10110"],
            segment=segment,
        )
        bank_to = BankAccount.objects.create(
            code="E2E-PNB", name="PNB E2E", account_type="checking",
            bank_name="PNB", bank_code="PNB", gl_account=accounts["10010"],
            segment=segment,
        )
        client.force_login(roles["head"])
        resp = client.post(
            "/cash/transfers/new/",
            {"from_account": bank_from.id, "to_account": bank_to.id,
             "amount": "25000.00", "purpose": "E2E fund move",
             "transfer_date": "2026-01-14"},
        )
        assert resp.status_code == 302
        transfer = InterAccountTransfer.objects.get()
        assert transfer.journal_entry.status == PostingStatus.POSTED

        # 8. Advance + liquidation --------------------------------------------
        from apps.ap.services import AdvanceService

        client.force_login(roles["head"])
        adv = AdvanceService.start(
            employee_name="E2E Officer", kind="officer", segment=segment,
            granted_date=date(2026, 1, 8), amount="20000.00", user=roles["head"],
        )
        resp = client.post(
            f"/ap/advances/{adv.id}/liquidate/",
            {"amount": "5000.00", "liquidate_date": "2026-01-15"},
        )
        assert resp.status_code == 302
        adv.refresh_from_db()
        assert adv.status == "partially_liquidated"
        assert adv.liquidated_amount == Decimal("5000.00")

        # 9. Weekly cycles + COLLECTIBLES + cash flow --------------------------
        client.force_login(roles["staff"])
        resp = client.post(
            "/cash/cycles/generate/",
            {"segment": segment.id, "start_date": "2026-01-06", "end_date": "2026-01-26"},
        )
        assert resp.status_code == 302
        assert WeeklyCashCycle.objects.count() == 3
        cycle = WeeklyCashCycle.objects.get(cycle_start=date(2026, 1, 6))
        CollectiblesService.generate(cycle)
        assert CollectiblesWorksheet.objects.filter(cycle=cycle).count() == 2
        cf = CashFlowService.generate(date(2026, 1, 6), date(2026, 1, 26), segment)
        assert cf.identity_holds

        # 10. Render all six register screens ---------------------------------
        client.force_login(user)
        screens = {
            "/journal/general/": [b"GENERAL JOURNAL", str(rfp.ap_number).encode(), b"OK"],
            f"/reports/cash-flow/?segment={segment.id}&period_start=2026-01-06&period_end=2026-01-26":
                [b"CASH FLOW STATEMENT", b"Identity"],
            "/cash/collectibles/?cycle=%d" % cycle.id: [b"COLLECTIBLES WORKSHEET", b"Distribution"],
            "/ar/aging/": [b"AR AGING", b"E2E Customer"],
            "/ap/advances/": [b"ADVANCES TO EMPLOYEES", b"E2E Officer", b"15,000.00"],
            "/cash/transfers/": [b"INTER-ACCOUNT TRANSFERS", b"E2E fund move"],
        }
        for url, needles in screens.items():
            resp = client.get(url)
            assert resp.status_code == 200, f"{url} -> {resp.status_code}"
            body = resp.content
            for needle in needles:
                assert needle in body, f"{url} missing {needle!r}"

        # 10b. Excel exports: all six endpoints return real, mirrored workbooks
        from io import BytesIO

        from openpyxl import load_workbook

        def export(url):
            resp = client.get(url)
            assert resp.status_code == 200, f"{url} -> {resp.status_code}"
            assert resp["Content-Type"] == (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            return load_workbook(BytesIO(resp.content))

        wb = export("/reports/trial-balance/export/?year=2026")
        ws = wb["TRIAL BALANCE"]
        assert "I2:AJ2" in {str(r) for r in ws.merged_cells.ranges}
        assert ws["K5"].value == "JANUARY"
        assert ws["A6"].value

        wb = export(
            "/reports/sfp/export/?period_start=2026-01-01&period_end=2026-01-31&net_profit=127000.00",
            )
        assert wb["YEAR END"]["B4"].value == "STATEMENT OF FINANCIAL POSITION"

        wb = export(
            "/reports/soce/export/?period_start=2026-01-01&period_end=2026-01-31&net_profit=127000.00",
            )
        assert wb["EQUITY"]["C2"].value == "STATEMENT OF CHANGES IN EQUITY"
        assert wb["EQUITY"]["D12"].value is not None

        wb = export("/reports/cos/export/?period_start=2026-01-01&period_end=2026-01-31")
        assert wb["COST OF SALES"]["A12"].value == "COGS - Fuel Purchase"
        assert wb["COST OF SALES"]["E54"].value is not None

        wb = export("/reports/te/export/?period_start=2026-01-01&period_end=2026-01-31",
                    )
        assert wb["January 2026 CGSE"]["H37"].value > 0

        wb = export(
            "/reports/cash-flow/export/?segment=%d&period_start=2026-01-06&period_end=2026-01-26"
            % segment.id)
        assert wb["CF"]["H5"].value == "Amounts in pesos"
        assert wb["CF"]["H25"].value is not None

        # 11. Immutability: posted entries cannot be reversed through the UI
        from apps.posting.models import JournalEntry

        posted = JournalEntry.objects.get(status=PostingStatus.POSTED, source_doc_no=rfp.ap_number)
        resp = client.post(f"/journal/{posted.id}/reverse/")
        assert resp.status_code == 302
        posted.refresh_from_db()
        assert posted.status == PostingStatus.POSTED

    def test_aging_buckets_reflect_open_invoices(self, client, company, segment, accounts, user):
        from apps.ar.models import Customer

        client.force_login(user)
        customer = Customer.objects.create(
            code="E2E-AG", name="E2E Ager", segment=segment
        )
        today = date.today()
        for offset, total in ((5, "10000.00"), (40, "20000.00"), (75, "30000.00")):
            create_invoice(segment, customer, today - timedelta(days=offset), total)
        resp = client.get("/ar/aging/")
        assert resp.status_code == 200
        body = resp.content
        assert b"0-30 days" in body and b"31-60 days" in body and b"61-90 days" in body
        assert b"E2E Ager" in body
        # Amounts render with thousand separators
        assert b"10,000.00" in body and b"20,000.00" in body and b"30,000.00" in body
        assert b"60,000.00" in body  # register total

    def test_pagination_pages_render(self, client, company, segment, accounts, user):
        """Every paginated screen honours ?page= (clamps out-of-range pages)."""
        from apps.ar.models import Customer

        client.force_login(user)
        for url in (
            "/journal/general/?page=2",
            "/foundation/coa/?page=2",
            "/ar/aging/?page=2",
            "/ap/advances/?page=2",
            "/cash/transfers/?page=2",
            "/journal/general/?page=999",
            "/foundation/coa/?page=999",
        ):
            resp = client.get(url)
            assert resp.status_code == 200, f"{url} -> {resp.status_code}"

    def test_coa_list_renders_and_filters(self, client, company, segment, accounts, user):
        client.force_login(user)
        resp = client.get("/foundation/coa/")
        assert resp.status_code == 200
        body = resp.content
        assert b"CHART OF ACCOUNTS" in body
        assert b"10010" in body and b"61100" in body and b"64110" in body
        # Segment filter narrows the list
        resp = client.get("/foundation/coa/?segment=DHPP")
        assert resp.status_code == 200
        # Name search
        resp = client.get("/foundation/coa/?q=Withholding")
        assert resp.status_code == 200
        assert b"64110" in resp.content
        assert b"61100" not in resp.content
        # Account-type filter
        resp = client.get("/foundation/coa/?account_type=liability")
        assert resp.status_code == 200
        assert b"64110" in resp.content
        assert b"Cash on Hand" not in resp.content
