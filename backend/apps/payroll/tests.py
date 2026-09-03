"""Payroll GL feed contract tests (BUILD-PLAN Phase 6, ADR-033).

Covers:
- ingest: line validation, balance enforcement, idempotency on batch ref
- preview: immutable lines shown, no editing
- post: one balanced immutable JE, batch-reference linkage, §14.1 lines
- reject: never posts
- parse_workbook: SUMMARY + JE LINES sheet to structured data
"""

from datetime import date
from decimal import Decimal

import pytest

from apps.foundation.models import Account, Company, Segment
from apps.posting.models import GeneralLedger, JournalEntry

from .models import PayrollFeed, PayrollFeedStatus

pytestmark = pytest.mark.django_db


@pytest.fixture
def company(db):
    return Company.objects.create(code="STMIET", name="Seven-Trent Machineries")


@pytest.fixture
def segment(db, company):
    return Segment.objects.create(code="DHPP", name="Diesel & Heavy Parts", company=company)


@pytest.fixture
def payroll_accounts(db):
    """COA slice used by the §14.1 payroll JE (gross, accrued, govt, WHT)."""
    rows = [
        ("63480", "Salary - R&F", "expense"),
        ("63470", "Overtime Pay", "expense"),
        ("22020", "Accrued Salaries", "liability"),
        ("23010", "SSS Payable (EE)", "liability"),
        ("23020", "PHIC Payable (EE)", "liability"),
        ("23030", "HDMF Payable (EE)", "liability"),
        ("64100", "Withholding Tax - Compensation", "liability"),
        ("61800", "Govt ER Shares", "expense"),
        ("23040", "SSS Payable (ER)", "liability"),
    ]
    out = {}
    for code, name, atype in rows:
        out[code] = Account.objects.create(
            code=code, name=name, account_type=atype, segment=Account.segment_for_code(code)
        )
    return out


def _jan_lines(payroll_accounts):
    """§14.1 gross-to-net: combos that produce a balanced batch."""
    return [
        {"line_no": 1, "gl_account": "63480", "debit": "80000.00", "credit": "0", "remarks": "Salary"},
        {"line_no": 2, "gl_account": "63470", "debit": "5000.00", "credit": "0", "remarks": "OT"},
        {"line_no": 3, "gl_account": "22020", "debit": "0", "credit": "60264.54", "remarks": "Net pay"},
        {"line_no": 4, "gl_account": "23010", "debit": "0", "credit": "8000.00", "remarks": "SSS EE"},
        {"line_no": 5, "gl_account": "23020", "debit": "0", "credit": "3000.00", "remarks": "PHIC EE"},
        {"line_no": 6, "gl_account": "23030", "debit": "0", "credit": "2235.46", "remarks": "HDMF EE"},
        {"line_no": 7, "gl_account": "64100", "debit": "0", "credit": "11500.00", "remarks": "WHT"},
    ]


@pytest.fixture
def batch(payroll_accounts, company, segment):
    from .services import PayrollFeedService

    lines = _jan_lines(payroll_accounts)
    feed = PayrollFeedService.ingest(
        batch_reference="PR-2026-08-A",
        period_start="2026-08-01", period_end="2026-08-15",
        entity="STMIET", lines=lines, company=company, segment_code="DHPP",
    )
    return feed


class TestIngest:
    def test_valid_batch_validated(self, payroll_accounts, company):
        from .services import PayrollFeedService

        feed = PayrollFeedService.ingest(
            batch_reference="PR-1", period_start="2026-08-01", period_end="2026-08-15",
            entity="STMIET", lines=_jan_lines(payroll_accounts), company=company,
        )
        assert feed.status == PayrollFeedStatus.VALIDATED
        assert feed.lines.count() == 7
        assert feed.net_pay_total > 0

    def test_unbalanced_raises(self, payroll_accounts, company):
        from apps.core.exceptions import ValidationError
        from .services import PayrollFeedService

        lines = _jan_lines(payroll_accounts)
        lines[2]["credit"] = "50000.00"  # breaks the balance
        with pytest.raises(ValidationError, match="balance"):
            PayrollFeedService.ingest(
                batch_reference="PR-BAD", period_start="2026-08-01", period_end="2026-08-15",
                entity="STMIET", lines=lines, company=company,
            )
        assert PayrollFeed.objects.filter(batch_reference="PR-BAD").count() == 0

    def test_unknown_gl_raises(self, payroll_accounts, company):
        from apps.core.exceptions import ValidationError
        from .services import PayrollFeedService

        lines = _jan_lines(payroll_accounts)
        lines[0]["gl_account"] = "99999"
        with pytest.raises(ValidationError, match="not found"):
            PayrollFeedService.ingest(
                batch_reference="PR-X", period_start="2026-08-01", period_end="2026-08-15",
                entity="STMIET", lines=lines, company=company,
            )

    def test_both_sides_rejected(self, payroll_accounts, company):
        from apps.core.exceptions import ValidationError
        from .services import PayrollFeedService

        lines = _jan_lines(payroll_accounts)
        lines[0]["credit"] = "1.00"  # both debit + credit on a line
        with pytest.raises(ValidationError):
            PayrollFeedService.ingest(
                batch_reference="PR-BS", period_start="2026-08-01", period_end="2026-08-15",
                entity="STMIET", lines=lines, company=company,
            )

    def test_idempotent_on_batch_ref(self, batch):
        from .services import PayrollFeedService

        again = PayrollFeedService.ingest(
            batch_reference="PR-2026-08-A", period_start="2026-08-01", period_end="2026-08-15",
            entity="STMIET", lines=[],
        )
        assert again == batch
        assert PayrollFeed.objects.filter(batch_reference="PR-2026-08-A").count() == 1


class TestPreviewAndPost:
    def test_preview_is_immutable_lines(self, batch):
        from .services import PayrollFeedService

        preview = PayrollFeedService.preview(batch)
        assert len(preview) == 7
        assert preview[0]["gl_account"] == "63480"
        assert preview[0]["debit"] == Decimal("80000.00")

    def test_post_creates_balanced_immutable_je(self, batch):
        from .services import PayrollFeedService

        feed = PayrollFeedService.post(batch)
        assert feed.status == PayrollFeedStatus.POSTED
        je = feed.journal_entry
        assert je is not None
        assert je.is_balanced and je.is_posted
        assert je.source_doc_no == "PR-2026-08-A"
        lines = {l.account.code: l for l in je.lines.all()}
        assert lines["63480"].debit == Decimal("80000.00")
        assert lines["22020"].credit == Decimal("60264.54")
        assert lines["64100"].credit == Decimal("11500.00")

    def test_post_creates_gl_rows(self, batch):
        from .services import PayrollFeedService

        feed = PayrollFeedService.post(batch)
        count = GeneralLedger.objects.filter(
            company=feed.journal_entry.company, transaction_date=feed.period_end
        ).count()
        assert count >= 7  # one GL row per JE line

    def test_reject_never_posts(self, batch):
        from .services import PayrollFeedService

        feed = PayrollFeedService.reject(batch, reason="Wrong net pay totals")
        assert feed.status == PayrollFeedStatus.REJECTED
        assert feed.journal_entry is None
        assert JournalEntry.objects.filter(source_doc_no="PR-2026-08-A").count() == 0


class TestParseWorkbook:
    def test_parse_xlsx(self, tmp_path):
        import openpyxl

        from .services import PayrollFeedService

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SUMMARY"
        ws.append(["batch_reference", "PR-2026-08-B"])
        ws.append(["period_start", "2026-08-01"])
        ws.append(["period_end", "2026-08-15"])
        ws.append(["entity", "STMIET"])
        ws.append(["segment", "DHPP"])
        ws.append(["cost_center", "AG"])
        ws2 = wb.create_sheet("JE LINES")
        ws2.append(["line_no", "segment", "gl_account", "debit", "credit", "remarks"])
        ws2.append([1, "DHPP", "63480", 1000.00, 0, "Salary"])
        ws2.append([2, "DHPP", "22020", 0, 1000.00, "Net pay"])
        path = str(tmp_path / "feed.xlsx")
        wb.save(path)

        with open(path, "rb") as f:
            data = PayrollFeedService.parse_workbook(f)
        assert data["batch_reference"] == "PR-2026-08-B"
        assert data["period_start"] == date(2026, 8, 1)
        assert len(data["lines"]) == 2
        assert data["lines"][0]["gl_account"] == "63480"

    def test_parse_and_ingest_xlsx(self, tmp_path, payroll_accounts, segment):
        import openpyxl

        from .services import PayrollFeedService

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SUMMARY"
        ws.append(["batch_reference", "PR-2026-08-C"])
        ws.append(["period_start", "2026-08-01"])
        ws.append(["period_end", "2026-08-15"])
        ws.append(["entity", "STMIET"])
        ws.append(["segment", "DHPP"])
        ws2 = wb.create_sheet("JE LINES")
        ws2.append(["line_no", "gl_account", "debit", "credit", "remarks"])
        ws2.append([1, "63480", 8000.00, 0, "Salary"])
        ws2.append([2, "22020", 0, 8000.00, "Net pay"])
        path = str(tmp_path / "feed2.xlsx")
        wb.save(path)

        with open(path, "rb") as f:
            data = PayrollFeedService.parse_workbook(f)
        feed = PayrollFeedService.ingest(**data)
        assert feed.status == PayrollFeedStatus.VALIDATED
        assert feed.lines.count() == 2


class TestHttpEndpoints:
    def test_ingest_and_post(self, api, payroll_accounts, segment):
        resp = api.post(
            "/api/v1/payroll/feeds/",
            {
                "batch_reference": "PR-HTTP", "period_start": "2026-08-01",
                "period_end": "2026-08-15", "entity": "STMIET", "segment": "DHPP",
                "lines": [
                    {"gl_account": "63480", "debit": "5000.00", "credit": "0"},
                    {"gl_account": "22020", "debit": "0", "credit": "5000.00"},
                ],
            },
            format="json",
        )
        assert resp.status_code == 201
        ref = resp.data["batch_reference"]

        detail = api.get(f"/api/v1/payroll/feeds/{ref}/")
        assert detail.status_code == 200
        assert len(detail.data["preview"]) == 2

        posted = api.post(f"/api/v1/payroll/feeds/{ref}/", {"action": "post"}, format="json")
        assert posted.status_code == 200
        assert posted.data["status"] == "posted"
        assert posted.data["journal_entry"]


@pytest.fixture
def api(client):
    from rest_framework.test import APIClient

    return APIClient()