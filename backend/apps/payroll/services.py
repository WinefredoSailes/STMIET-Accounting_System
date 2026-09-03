"""Payroll GL feed services (BUILD-PLAN Phase 6, ADR-033).

Pipeline: ingest feed → validate schema/master-data → JE preview (immutable) →
reviewer approve → post to immutable journal. Corrections are NOT made here —
they go back to the payroll system (single source of truth). The feed file is
archived with the batch for audit.

Validation (ADR-033 §2):
- every line has a postable COA GL account code;
- the batch's lines balance (sum(debit) == sum(credit));
- amounts are 2-dp money;
- entity/segment/cost-center resolve (segment optional for ALL).
"""

import mimetypes
from datetime import date, datetime
from decimal import Decimal

from django.db import transaction
from django.utils import timezone

from apps.core.exceptions import ValidationError
from apps.core.money import money
from apps.foundation.models import Account, Company, Segment
from apps.posting.models import JournalEntry, JournalEntryLine, PostingStatus
from apps.posting.services import PostingService

from .models import PayrollFeed, PayrollFeedLine, PayrollFeedStatus


class PayrollFeedService:
    """Stateless importer for the payroll GL feed contract."""

    REQUIRED_SUMMARY_FIELDS = {
        "period_start", "period_end", "entity", "batch_reference",
    }

    # ------------------------------------------------------------------ ingest

    @classmethod
    def ingest(cls, *, batch_reference, period_start, period_end, entity,
               lines, segment_code=None, cost_center="", schema_version="v1",
               archived_file=None, user=None, company=None) -> PayrollFeed:
        """Validate and persist one payroll feed batch (idempotent on reference).

        `lines` is an iterable of dicts:
            {line_no, segment, cost_center, gl_account, debit, credit, remarks}
        A previously-persisted batch_reference returns the existing feed (no-op).
        """
        existing = PayrollFeed.objects.filter(batch_reference=batch_reference).first()
        if existing:
            return existing

        segment = cls._resolve_segment(segment_code) if segment_code else None
        if company is None:
            company = segment.company if segment else Company.objects.first()
        if company is None:
            raise ValidationError("Payroll feed requires a company.")

        feed = PayrollFeed(
            batch_reference=batch_reference,
            schema_version=schema_version,
            period_start=_coerce_date(period_start),
            period_end=_coerce_date(period_end),
            entity=entity,
            company=company,
            segment=segment,
            cost_center=cost_center,
            archived_file=archived_file or "",
            status=PayrollFeedStatus.UPLOADED,
            created_by=user,
        )

        parsed_lines = [cls._validate_line(line, segment) for line in lines]
        feed.net_pay_total = money(sum(l["debit"] for l in parsed_lines if l["net_leg"]))
        # balance check across Debit + Credit legs
        total_debit = sum(l["debit"] for l in parsed_lines)
        total_credit = sum(l["credit"] for l in parsed_lines)
        if total_debit != total_credit:
            raise ValidationError(
                f"Payroll batch '{batch_reference}' does not balance: "
                f"debit {total_debit} != credit {total_credit}."
            )

        with transaction.atomic():
            feed.save()
            for i, line in enumerate(parsed_lines, start=1):
                PayrollFeedLine.objects.create(
                    feed=feed, line_no=line["line_no"] or i,
                    segment=line["segment"], cost_center=line["cost_center"],
                    gl_account=line["gl_account"], debit=line["debit"],
                    credit=line["credit"], remarks=line["remarks"],
                )
        feed.status = PayrollFeedStatus.VALIDATED
        feed.save(update_fields=["status"])
        return feed

    @classmethod
    def _validate_line(cls, raw, default_segment) -> dict:
        try:
            gl_acct = Account.objects.get(code=raw["gl_account"], is_postable=True)
        except Account.DoesNotExist:
            raise ValidationError(f"Payroll GL account '{raw.get('gl_account')}' not found in COA.")
        except KeyError:
            raise ValidationError("Payroll line missing gl_account.")

        debit = money(raw.get("debit") or 0)
        credit = money(raw.get("credit") or 0)
        if debit < 0 or credit < 0:
            raise ValidationError("Payroll line amounts must be non-negative.")
        if debit and credit:
            raise ValidationError("Payroll line has both debit and credit (must be one or the other).")
        if not debit and not credit:
            raise ValidationError("Payroll line has neither debit nor credit.")

        segment = default_segment
        seg_code = raw.get("segment")
        if seg_code:
            segment = cls._resolve_segment(seg_code)

        return {
            "line_no": raw.get("line_no"),
            "segment": segment,
            "cost_center": raw.get("cost_center", ""),
            "gl_account": gl_acct,
            "debit": debit,
            "credit": credit,
            "remarks": raw.get("remarks", ""),
            "net_leg": debit > 0,
        }

    @classmethod
    def _resolve_segment(cls, code):
        try:
            return Segment.objects.get(code=code)
        except Segment.DoesNotExist:
            raise ValidationError(f"Segment '{code}' not found.")

    # ------------------------------------------------------------------ review

    @classmethod
    def preview(cls, feed: PayrollFeed) -> list:
        """Immutable JE preview: the exact lines that will post (no editing)."""
        if feed.status != PayrollFeedStatus.VALIDATED:
            raise ValidationError("Feed must be VALIDATED before preview/review.")
        return [
            {
                "line_no": l.line_no,
                "segment": l.segment.code if l.segment else feed.entity,
                "cost_center": l.cost_center,
                "gl_account": l.gl_account.code,
                "account_name": l.gl_account.name,
                "debit": l.debit,
                "credit": l.credit,
                "remarks": l.remarks,
            }
            for l in feed.lines.order_by("line_no")
        ]

    @classmethod
    def post(cls, feed: PayrollFeed, *, user=None) -> PayrollFeed:
        """Reviewer approves: post the feed's lines as one immutable JE."""
        if feed.status == PayrollFeedStatus.POSTED:
            return feed
        if feed.status not in (PayrollFeedStatus.VALIDATED, PayrollFeedStatus.REVIEWED):
            raise ValidationError("Feed must be validated before posting.")
        if feed.segment is None:
            # ADR-011: a JE must belong to exactly one segment. An "ALL" feed
            # spans segments; v1 does not invent an allocation — the producer
            # must emit per-segment feeds.
            raise ValidationError(
                "Feed has no segment. Posting requires a single segment "
                "(ADR-011); emit per-segment payroll feeds."
            )

        company = feed.company
        journal = JournalEntry(
            entry_no=f"PAY-{feed.batch_reference}",
            company=company,
            segment=feed.segment,
            transaction_date=feed.period_end,
            status=PostingStatus.DRAFT,
            description=f"Payroll {feed.period_start} to {feed.period_end} ({feed.entity})",
            source_doc_type="PAY",
            source_doc_no=feed.batch_reference,
            created_by=user,
        )
        journal.save()
        with transaction.atomic():
            for i, line in enumerate(feed.lines.order_by("line_no"), start=1):
                JournalEntryLine.objects.create(
                    entry=journal, line_no=i, account=line.gl_account,
                    debit=line.debit, credit=line.credit, description=line.remarks,
                )
            journal.recalc_totals()
            PostingService.post(journal, user=user)

        feed.journal_entry = journal
        feed.status = PayrollFeedStatus.POSTED
        feed.review_user = user
        feed.reviewed_at = timezone.now()
        feed.posted_at = timezone.now()
        feed.save(update_fields=["journal_entry", "status", "review_user", "reviewed_at", "posted_at"])
        return feed

    @classmethod
    def reject(cls, feed: PayrollFeed, *, reason: str = "", user=None) -> PayrollFeed:
        """Reject the batch — it never posts; a corrected feed replaces it."""
        if feed.status == PayrollFeedStatus.POSTED:
            raise ValidationError("Posted feeds cannot be rejected.")
        feed.status = PayrollFeedStatus.REJECTED
        feed.validation_error = reason[:1000]
        feed.review_user = user
        feed.reviewed_at = timezone.now()
        feed.save(update_fields=["status", "validation_error", "review_user", "reviewed_at"])
        return feed

    # ------------------------------------------------------------------ parse

    @classmethod
    def parse_workbook(cls, f) -> dict:
        """Read a .xlsx/.csv payroll feed into {summary, lines} for ingest.

        Default workbook layout mirrors ADR-033: sheet 'SUMMARY' carries the
        batch header; sheet 'JE LINES' carries one row per GL line.
        """
        name = getattr(f, "name", "") or ""
        if name.lower().endswith(".csv"):
            rows = cls._parse_csv(f)
        else:
            rows = cls._parse_xlsx(f)
        summary = rows["summary"]
        lines = rows["lines"]
        try:
            period_start = _coerce_date(summary["period_start"])
            period_end = _coerce_date(summary["period_end"])
        except KeyError as exc:
            raise ValidationError(f"SUMMARY missing required field: {exc.args[0]}") from exc

        cleaned_lines = []
        for row in lines:
            debit = _num(row.get("debit"))
            credit = _num(row.get("credit"))
            cleaned_lines.append(
                {
                    "line_no": row.get("line_no"),
                    "segment": row.get("segment"),
                    "cost_center": row.get("cost_center", ""),
                    "gl_account": str(row.get("gl_account")),
                    "debit": debit,
                    "credit": credit,
                    "remarks": str(row.get("remarks", "") or ""),
                }
            )
        return {
            "batch_reference": summary["batch_reference"],
            "period_start": period_start,
            "period_end": period_end,
            "entity": summary["entity"],
            "segment_code": summary.get("segment"),
            "cost_center": summary.get("cost_center", ""),
            "lines": cleaned_lines,
        }

    @classmethod
    def _parse_xlsx(cls, f):
        import openpyxl

        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        summary_ws = wb["SUMMARY"] if "SUMMARY" in wb.sheetnames else wb[wb.sheetnames[0]]
        summary = {}
        for row in summary_ws.iter_rows(values_only=True):
            if row and row[0] is not None:
                key = str(row[0]).strip()
                val = row[1] if len(row) > 1 else None
                summary[key] = val

        lines_ws = wb["JE LINES"] if "JE LINES" in wb.sheetnames else None
        lines = []
        if lines_ws is not None:
            rows = list(lines_ws.iter_rows(values_only=True))
            if rows:
                header = [str(c).strip().lower() if c else "" for c in rows[0]]
                for row in rows[1:]:
                    if not row or all(v is None for v in row):
                        continue
                    r = {header[i]: row[i] for i in range(len(header)) if i < len(row)}
                    lines.append(r)
        return {"summary": summary, "lines": lines}

    @classmethod
    def _parse_csv(cls, f):
        import csv
        import io

        text = f.read().decode("utf-8-sig")
        data = list(csv.DictReader(io.StringIO(text)))
        # First row(s) = summary, remainder = lines (best-effort leniency).
        summary = {}
        lines = []
        for row in data:
            flat = {k.strip().lower().replace(" ", "_"): v for k, v in row.items() if k}
            if "batch_reference" in flat or "period_start" in flat:
                summary.update(flat)
            else:
                lines.append(_norm(flat))
        return {"summary": summary, "lines": lines}


def _norm(row):
    return {
        "line_no": row.get("line_no"),
        "segment": row.get("segment"),
        "cost_center": row.get("cost_center", ""),
        "gl_account": row.get("gl_account"),
        "debit": _num(row.get("debit")),
        "credit": _num(row.get("credit")),
        "remarks": row.get("remarks", ""),
    }


def _num(v):
    if v in (None, ""):
        return Decimal("0.00")
    return money(v)


def _coerce_date(v):
    """Accept date, datetime, or a yyyy-mm-dd string."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return date.fromisoformat(str(v).strip())