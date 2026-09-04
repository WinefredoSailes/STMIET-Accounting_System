"""Bulk-import the bank account master (BUILD-PLAN Phase 10 master-data migration).

Usage:
    py manage.py import_banks --file excel-files/BANKS.xlsx
    py manage.py import_banks --file excel-files/BANKS.csv

Expected columns (locate by header name):

    CODE | NAME | TYPE | BANK NAME | BANK CODE | GL ACCOUNT | ADB REQUIRED

  - TYPE         : checking | savings | pcf_coh (defaults checking)
  - GL ACCOUNT   : COA 5-digit code (e.g. 10040) — must exist (run import_coa)
  - ADB REQUIRED : maintaining balance (defaults 5000.00; PCF/COH 0)

Idempotent: a CODE that exists is updated in place; re-runs are safe.
"""

import csv
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.cash.models import BankAccount, BankAccountType
from apps.core.money import money
from apps.foundation.models import Account, Company

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

HEADER_MAP = {
    "code": ("CODE", "ACCOUNT CODE", "BANK CODE", "ACCT #"),
    "name": ("NAME", "ACCOUNT NAME", "BANK ACCOUNT"),
    "type": ("TYPE", "ACCOUNT TYPE", "KIND"),
    "bank_name": ("BANK NAME",),
    "bank_code": ("BANK CODE", "BANK"),
    "gl": ("GL ACCOUNT", "GL", "COA", "ACCOUNT"),
    "adb": ("ADB REQUIRED", "ADB", "AVERAGE DAILY BALANCE", "MINIMUM BALANCE"),
}

TYPE_ALIASES = {
    "CHECKING": BankAccountType.CHECKING, "CURRENT": BankAccountType.CHECKING,
    "SAVINGS": BankAccountType.SAVINGS, "SAVING": BankAccountType.SAVINGS,
    "PCF": BankAccountType.PCF_COH, "PCF_COH": BankAccountType.PCF_COH,
    "PETTY CASH": BankAccountType.PCF_COH, "CASH ON HAND": BankAccountType.PCF_COH,
    "COH": BankAccountType.PCF_COH,
}


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Bulk-import the bank account master from CSV or XLSX (idempotent)."

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file", default=None, help="CSV/XLSX path")
        parser.add_argument("--company", dest="company", default="STMIET")

    def _load_rows(self, file_path):
        if file_path.suffix.lower() == ".csv":
            with open(file_path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.reader(fh)
                header = [h.strip().upper() for h in next(reader, [])]
                for i, row in enumerate(reader, start=2):
                    if not any(row):
                        continue
                    yield self._map_row(row, header), f"CSV row {i}"
        else:
            if openpyxl is None:
                raise CommandError("openpyxl required for xlsx input; pip install openpyxl")
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header = None
                header_idx = 0
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if not row or not any(row):
                        continue
                    upper = [str(v or "").strip().upper() for v in row]
                    if header is None and any("CODE" in h or "NAME" in h for h in upper):
                        header = upper
                        header_idx = i
                        break
                if header is None:
                    self.stdout.write(self.style.WARNING(
                        f"skip sheet {sheet_name}: no recognizable header row"
                    ))
                    continue
                for i, row in enumerate(ws.iter_rows(min_row=header_idx + 1, values_only=True), start=header_idx + 1):
                    if not row or not any(row):
                        continue
                    yield self._map_row(row, header), f"{sheet_name} row {i}"

    def _map_row(self, row, header):
        idx = {}
        for field, names in HEADER_MAP.items():
            for n in names:
                if n in header:
                    idx[field] = header.index(n)
                    break
        out = {}
        for field, pos in idx.items():
            if pos < len(row):
                out[field] = _clean(row[pos])
        return out

    def _resolve_type(self, raw):
        if not raw:
            return BankAccountType.CHECKING
        return TYPE_ALIASES.get(raw.upper().strip(), BankAccountType.CHECKING)

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = Path(options["file"]) if options["file"] else None
        if file_path is None or not file_path.exists():
            repo_root = Path(__file__).resolve().parents[5]
            for cand in (
                repo_root / "excel-files" / "BANKS.xlsx",
                repo_root / "excel-files" / "BANKS.csv",
                Path.cwd() / "excel-files" / "BANKS.xlsx",
            ):
                if cand.exists():
                    file_path = cand
                    break
        if file_path is None or not file_path.exists():
            raise CommandError(
                "Bank master not found. Pass --file, or place BANKS.csv/.xlsx"
                " under /excel-files."
            )

        company = Company.objects.filter(code=options["company"]).first()
        if company is None:
            raise CommandError(f"Company '{options['company']}' not found. Run import_coa first.")

        created = updated = skipped = 0
        for row, label in self._load_rows(file_path):
            code = row.get("code", "")
            name = row.get("name", "")
            if not code and not name:
                skipped += 1
                continue
            try:
                account_type = self._resolve_type(row.get("type", ""))
                gl_code = row.get("gl", "")
                if gl_code:
                    gl_account = Account.objects.filter(code=gl_code, is_postable=True).first()
                    if gl_account is None:
                        self.stdout.write(self.style.ERROR(
                            f"  {label}: COA account {gl_code} missing. Run import_coa first."
                        ))
                        skipped += 1
                        continue
                else:
                    self.stdout.write(self.style.ERROR(f"  {label}: no GL account provided."))
                    skipped += 1
                    continue

                adb_raw = row.get("adb", "")
                adb = money(adb_raw) if adb_raw else Decimal("0.00" if account_type == BankAccountType.PCF_COH else "5000.00")

                existing_gl = BankAccount.objects.filter(gl_account=gl_account).first()
                if existing_gl is not None and existing_gl.code != code:
                    self.stdout.write(self.style.WARNING(
                        f"  {label}: GL account {gl_code} already used by "
                        f"bank {existing_gl.code} — skipped (account must be unique)."
                    ))
                    skipped += 1
                    continue

                _, was_created = BankAccount.objects.update_or_create(
                    code=code,
                    defaults={
                        "name": name or code,
                        "account_type": account_type,
                        "bank_name": row.get("bank_name", ""),
                        "bank_code": row.get("bank_code", ""),
                        "gl_account": gl_account,
                        "company": company,
                        "adb_required": adb,
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except (CommandError, ValueError) as exc:
                self.stdout.write(self.style.ERROR(f"  {label}: {exc}"))
                skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f"import_banks: {created} created, {updated} updated, {skipped} skipped."
        ))
