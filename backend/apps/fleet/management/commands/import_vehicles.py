"""Bulk-import the vehicle register (BUILD-PLAN Phase 10 master-data migration).

Usage:
    py manage.py import_vehicles --file excel-files/VEHICLES.xlsx
    py manage.py import_vehicles --file excel-files/VEHICLES.csv

Expected columns (locate by header name):

    PLATE NO | MAKE MODEL | TYPE | SEGMENT | ASSET NO

  - TYPE      : fuel_tanker | boom_truck | office_vehicle | other (defaults other)
  - SEGMENT   : "DHPP", "DMIE", "OPS" (optional)
  - ASSET NO  : links the vehicle to an existing Asset (optional; ADR-034)

Idempotent: a PLATE NO that exists is updated in place; re-runs are safe.
"""

import csv
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.assets.models import Asset
from apps.fleet.models import Vehicle, VehicleType
from apps.foundation.models import Segment

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

HEADER_MAP = {
    "plate": ("PLATE NO", "PLATE NUMBER", "PLATE", "VEHICLE NO", "UNIT NO"),
    "make": ("MAKE MODEL", "MAKE", "MODEL", "DESCRIPTION"),
    "vtype": ("TYPE", "VEHICLE TYPE", "KIND"),
    "segment": ("SEGMENT", "SECTION", "DEPARTMENT"),
    "asset_no": ("ASSET NO", "FA NO", "ASSET NUMBER", "LINKED ASSET"),
}

TYPE_ALIASES = {
    "FUEL TANKER": VehicleType.FUEL_TANKER, "TANKER": VehicleType.FUEL_TANKER,
    "BOOM TRUCK": VehicleType.BOOM_TRUCK, "BOOM": VehicleType.BOOM_TRUCK, "CRANE": VehicleType.BOOM_TRUCK,
    "OFFICE VEHICLE": VehicleType.OFFICE_VEHICLE, "CAR": VehicleType.OFFICE_VEHICLE,
    "SEDAN": VehicleType.OFFICE_VEHICLE, "TRUCK": VehicleType.OFFICE_VEHICLE,
    "OTHER": VehicleType.OTHER,
}

SEGMENT_ALIASES = {
    "DHPP": "DHPP", "DMIE": "DMIE", "OPS": "OPS", "": None,
}


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Bulk-import the vehicle register from CSV or XLSX (idempotent)."

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file", default=None, help="CSV/XLSX path")

    def _load_rows(self, file_path):
        if file_path.suffix.lower() == ".csv":
            with open(file_path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.reader(fh)
                header = [h.strip().upper() for h in next(reader, [])]
                for i, row in enumerate(reader, start=2):
                    if not any(row):
                        continue
                    yield self._map_row(row, header), f"CSV row {i}"
        else:
            if openpyxl is None:
                raise CommandError("openpyxl required for xlsx input; pip install openpyxl")
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header = None
                header_idx = 0
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if not row or not any(row):
                        continue
                    upper = [str(v or "").strip().upper() for v in row]
                    if header is None and any("PLATE" in h for h in upper):
                        header = upper
                        header_idx = i
                        break
                if header is None:
                    self.stdout.write(self.style.WARNING(
                        f"skip sheet {sheet_name}: no PLATE header"
                    ))
                    continue
                for i, row in enumerate(ws.iter_rows(min_row=header_idx + 1, values_only=True), start=header_idx + 1):
                    if not row or not any(row):
                        continue
                    yield self._map_row(row, header), f"{sheet_name} row {i}"

    def _map_row(self, row, header):
        idx = {}
        for field, names in HEADER_MAP.items():
            for n in names:
                if n in header:
                    idx[field] = header.index(n)
                    break
        out = {}
        for field, pos in idx.items():
            if pos < len(row):
                out[field] = _clean(row[pos])
        return out

    def _resolve_type(self, raw):
        if not raw:
            return VehicleType.OTHER
        return TYPE_ALIASES.get(raw.upper().strip(), VehicleType.OTHER)

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = Path(options["file"]) if options["file"] else None
        if file_path is None or not file_path.exists():
            repo_root = Path(__file__).resolve().parents[5]
            for cand in (
                repo_root / "excel-files" / "VEHICLES.xlsx",
                repo_root / "excel-files" / "VEHICLES.csv",
                Path.cwd() / "excel-files" / "VEHICLES.xlsx",
            ):
                if cand.exists():
                    file_path = cand
                    break
        if file_path is None or not file_path.exists():
            raise CommandError(
                "Vehicle master not found. Pass --file, or place VEHICLES.csv"
                "/.xlsx under /excel-files."
            )

        created = updated = skipped = 0
        for row, label in self._load_rows(file_path):
            plate = row.get("plate", "").upper().strip()
            if not plate:
                skipped += 1
                continue
            try:
                seg_code = SEGMENT_ALIASES.get(row.get("segment", "").upper().strip())
                segment = Segment.objects.filter(code=seg_code).first() if seg_code else None

                asset_no = row.get("asset_no", "")
                linked_asset = None
                if asset_no:
                    linked_asset = Asset.objects.filter(asset_no=asset_no).first()
                    if linked_asset is None:
                        self.stdout.write(self.style.WARNING(
                            f"  {label}: asset {asset_no} not found; vehicle created unlinked."
                        ))

                vehicle, was_created = Vehicle.objects.update_or_create(
                    plate_no=plate,
                    defaults={
                        "make_model": row.get("make", ""),
                        "vehicle_type": self._resolve_type(row.get("vtype", "")),
                        "segment": segment,
                        "is_active": True,
                    },
                )
                if linked_asset is not None:
                    linked_asset.vehicle = vehicle
                    linked_asset.save(update_fields=["vehicle", "updated_at"])
                if was_created:
                    created += 1
                else:
                    updated += 1
            except (CommandError, ValueError) as exc:
                self.stdout.write(self.style.ERROR(f"  {label}: {exc}"))
                skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f"import_vehicles: {created} created, {updated} updated, {skipped} skipped."
        ))
