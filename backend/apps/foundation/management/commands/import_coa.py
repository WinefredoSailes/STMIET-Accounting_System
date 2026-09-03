"""Seed/import base COA, companies, segments and fiscal calendar.

Usage:
    py manage.py import_coa --file excel-files/CHART-OF-ACCOUNTS_REVISED-SEPT-2026.xlsx
    py manage.py import_coa --company STMIET --fiscal-year 2026

The revised source workbook is split across FOUR sheets:
    * COA BALANCE SHEET   — assets / liabilities / equity (company-wide -> shared)
    * COA REV             — revenue (per-segment where applicable)
    * COA COS             — cost of sales (per-segment)
    * COA OPEX & NON OPEX — operating & non-operating expenses (shared)

The layout is NOT uniform across sheets, so this importer locates the
"REQUIRED SEGMENT" column by HEADER NAME rather than a fixed index. A sheet
without that column (the balance sheet) stores every account as SHARED.
Multi-segment values (e.g. "DHPP, DMIE, OPS") collapse to SHARED.

Account type is derived from the code prefix, which the workbook's own
"Major Accounts" column does not express reliably:
    1=Asset 2=Liability 3=Equity 4=Revenue 5/6=Expense
"""

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.foundation.models import Account, AccountType, Company, FiscalYear, Segment, SEGMENT_CHOICES

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

# Valid single-segment values in the workbook (others -> SHARED/ALL).
SINGLE_SEGMENTS = {c for c, _ in SEGMENT_CHOICES if c != "ALL"}
SHARED = "ALL"

# Standard column names used across the sheets.
HEADER = ("COA", "ACCOUNT TITLES")
REQUIRED_SEGMENT_COL = "REQUIRED SEGMENT"

# Prefix -> account type.
PREFIX_TYPE = {
    "1": AccountType.ASSET,
    "2": AccountType.LIABILITY,
    "3": AccountType.EQUITY,
    "4": AccountType.REVENUE,
    "5": AccountType.EXPENSE,
    "6": AccountType.EXPENSE,
}


class Command(BaseCommand):
    help = "Seed companies, segments, fiscal calendar and the chart of accounts."

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file", default=None, help="COA xlsx path")
        parser.add_argument("--company", dest="company", default="STMIET")
        parser.add_argument("--fiscal-year", dest="fiscal_year", default="2026")

    def _load_workbook(self, options):
        file_path = Path(options["file"]) if options["file"] else None
        if file_path is None:
            repo_root = Path(__file__).resolve().parents[5]
            for cand in (
                repo_root / "excel-files" / "CHART-OF-ACCOUNTS_REVISED-SEPT-2026.xlsx",
                repo_root / "excel-files" / "COA-STMIET-2026.xlsx",
                Path.cwd() / "excel-files" / "CHART-OF-ACCOUNTS_REVISED-SEPT-2026.xlsx",
            ):
                if cand.exists():
                    file_path = cand
                    break
        if file_path is None or not file_path.exists():
            raise CommandError(
                "COA workbook not found. Pass --file, or place the revised "
                "workbook in /excel-files."
            )
        if openpyxl is None:
            raise CommandError("openpyxl required; install with: pip install openpyxl")
        self.stdout.write(f"Reading {file_path.name}")
        return openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    def _iter_sheets(self, wb):
        """Yield (sheet_name, header_indexes, rows) per data sheet."""
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            header_idx = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if not row:
                    continue
                first = str(row[0] or "").strip().upper()
                second = str(row[1] or "").strip().upper() if len(row) > 1 else ""
                if first == "COA" and second == "ACCOUNT TITLES":
                    header_idx = i
                    break
                rows.append(row)  # pre-header rows are padding (ignored)
            if header_idx is None:
                continue  # not a COA data sheet
            yield name, header_idx, ws

    @transaction.atomic
    def handle(self, *args, **options):
        company, _ = Company.objects.get_or_create(
            code=options["company"],
            defaults={"name": "Seven-Trent Machineries Industrial Equipment Trading"},
        )
        segments = [
            ("DHPP", "Diesel & Heavy Parts Procurement", 0, company),
            ("DMIE", "Diesel Machinery & Industrial Equipment", 3, company),
            ("OPS", "Operations / Services", 6, company),
        ]
        for code, name, key, comp in segments:
            Segment.objects.get_or_create(
                code=code, defaults={"name": name, "coa_key_digit": key, "company": comp}
            )

        fy, _ = FiscalYear.objects.get_or_create(
            company=company,
            code=options["fiscal_year"],
            defaults={
                "start_date": f"{options['fiscal_year']}-01-01",
                "end_date": f"{options['fiscal_year']}-12-31",
            },
        )

        wb = self._load_workbook(options)
        imported = 0
        for sheet_name, header_idx, ws in self._iter_sheets(wb):
            header_row = next(
                ws.iter_rows(min_row=header_idx + 1, max_row=header_idx + 1, values_only=True)
            )
            col = {
                str(v).strip().upper(): i for i, v in enumerate(header_row) if v is not None
            }
            has_segment = REQUIRED_SEGMENT_COL in col

            def col_idx(*names):
                for n in names:
                    if n in col:
                        return col[n]
                return None

            idx_code = col_idx("COA")
            idx_name = col_idx("ACCOUNT TITLES")
            idx_class = col_idx("CLASSIFICATION")
            idx_cat = col_idx("CATEGORY")
            idx_sub = col_idx("SUB-ACCOUNTS", "SUB ACCOUNTS")
            idx_major = col_idx("MAJOR ACCOUNTS")
            idx_seg = col.get(REQUIRED_SEGMENT_COL)

            for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
                if not row:
                    continue
                code_val = row[idx_code] if idx_code is not None and idx_code < len(row) else None
                if code_val is None:
                    continue
                code = str(code_val).strip()
                if len(code) < 4 or not code.isdigit():
                    continue
                name = str(row[idx_name]).strip() if idx_name is not None else ""
                segment = self._resolve_segment(row, has_segment, idx_seg, code)
                prefix = code[0]
                atype = PREFIX_TYPE.get(prefix, AccountType.ASSET)
                classification = str(row[idx_class]).strip() if (idx_class is not None and idx_class < len(row)) else ""
                category = str(row[idx_cat]).strip() if (idx_cat is not None and idx_cat < len(row)) else ""
                sub_accounts = str(row[idx_sub]).strip() if (idx_sub is not None and idx_sub < len(row)) else ""
                major_accounts = str(row[idx_major]).strip() if (idx_major is not None and idx_major < len(row)) else ""

                Account.objects.update_or_create(
                    code=code,
                    defaults={
                        "name": name,
                        "account_type": atype,
                        "segment": segment,
                        "description": f"{classification} / {category}".strip(" /"),
                        "classification": classification,
                        "category": category,
                        "sub_accounts": sub_accounts,
                        "major_accounts": major_accounts,
                        "normal_balance": "debit"
                        if atype in (AccountType.ASSET, AccountType.EXPENSE)
                        else "credit",
                    },
                )
                acct_count = imported
        seeded_maps = self._seed_account_maps(company)
        self.stdout.write(
            self.style.SUCCESS(
                f"Imported {acct_count} accounts from {len(list(wb.sheetnames))} sheets "
                f"and seeded {seeded_maps} segment account maps."
            )
        )

    def _seed_account_maps(self, company) -> int:
        """Create the data-driven SegmentAccountMap rows (Phase 2) that replace
        the per-segment COA dicts in ap/assets. Codes below mirror the revised
        shared COA; the authoritative rows become DB data after this runs, so
        services never hardcode COA codes."""
        from apps.foundation.models import Account, Segment, SegmentAccountMap

        role_codes = {
            SegmentAccountMap.ROLE_AP: "20000",
            SegmentAccountMap.ROLE_AP_WHT: "64110",
            SegmentAccountMap.ROLE_CASH: "10010",
            SegmentAccountMap.ROLE_LOANS: "27010",
            SegmentAccountMap.ROLE_DISPOSAL_GAIN: "43070",
            SegmentAccountMap.ROLE_DISPOSAL_LOSS: "62000",
        }
        count = 0
        for seg in Segment.objects.filter(company=company):
            for role, code in role_codes.items():
                account = Account.objects.filter(code=code).first()
                if account is None:
                    self.stdout.write(
                        self.style.WARNING(f"skip {seg.code}:{role} -> COA {code} missing")
                    )
                    continue
                SegmentAccountMap.objects.update_or_create(
                    segment=seg, role=role, defaults={"account": account}
                )
                count += 1
        return count

    def _resolve_segment(self, row, has_segment: bool, idx_seg, code: str) -> str:
        """Return SHARED unless the sheet has a REQUIRED SEGMENT column whose
        value is exactly one segment. Balance-sheet rows (no such column) and
        multi-segment values collapse to SHARED."""
        if not has_segment or idx_seg is None or idx_seg >= len(row):
            return SHARED
        raw = row[idx_seg]
        if raw is None:
            return SHARED
        value = str(raw).strip().upper().replace(" ", "")
        if value in SINGLE_SEGMENTS:
            return value
        return SHARED
