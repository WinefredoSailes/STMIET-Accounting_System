"""Post opening balances from a Trial-Balance-style workbook (Phase 10).

Usage:
    py manage.py import_opening_balances --file excel-files/OPENING-BALANCES.xlsx
    py manage.py import_opening_balances --file excel-files/OPENING-BALANCES.csv
    py manage.py import_opening_balances --as-of 2026-01-01

Reads account balances (per COA code, per segment) and posts them as
balanced opening journal entries. Because the GL attributes every line of a
JE to that JE's segment (ADR-011), the importer creates ONE journal entry per
segment so per-segment Trial Balance and Financial Statement columns are
correct:

    Dr [asset/expense accounts] | Cr [liability/equity/revenue accounts]

  * Each segment JE (DHPP / DMIE / OPS / ALL) carries the balances for the
    accounts tagged to that segment.
  * The plug account for each segment defaults to that segment's
    `opening_equity` mapping (SegmentAccountMap, seeded by import_coa). When
    the Dr/Cr sides do not balance, the difference is posted to the plug
    account so the entry always balances — the plug is explicit and visible,
    never hidden.

Expected source layout (matching the TRIAL-BALANCE.xlsx workbook):

    COA | NORMAL BALANCE | ACCOUNT TITLES | SEGMENT | ... | OPENING DR | OPENING CR

Columns:
  * COA        : 5-digit account code (required — run import_coa first)
  * SEGMENT    : DHPP / DMIE / OPS / — (defaults to account.segment)
  * OPENING DR : debit opening balance
  * OPENING CR : credit opening balance

If the workbook uses different column headers, pass --dr-col and --cr-col.

Idempotent: posting checks whether an entry with the same (entry_no_prefix,
segment) already exists and skips it, so re-running is safe. A posted entry
can be reversed with the normal UI reversal.
"""

import csv
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.core.exceptions import AccountingError
from apps.core.money import money
from apps.foundation.models import (
    Account,
    Company,
    FiscalPeriod,
    Segment,
    SEGMENT_CHOICES,
    resolve_segment_account,
)
from apps.foundation.models import SegmentAccountMap
from apps.posting.models import JournalEntry, JournalEntryLine, PostingStatus
from apps.posting.services import PostingService

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

SINGLE_SEGMENTS = {c for c, _ in SEGMENT_CHOICES if c != "ALL"}
SHARED = "ALL"

HEADER_MAP = {
    "code": ("COA", "ACCOUNT CODE", "ACCT", "GL CODE"),
    "segment": ("SEGMENT", "SECTION", "DEPARTMENT"),
    "dr": ("OPENING DR", "DR", "DEBIT", "DEBITS", "DR AMOUNT"),
    "cr": ("OPENING CR", "CR", "CREDIT", "CREDITS", "CR AMOUNT"),
    "name": ("ACCOUNT TITLES", "ACCOUNT NAME", "NAME"),
}


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _to_decimal(value):
    if value is None or value == "":
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).replace(",", "").strip()
    text = text.replace("(", "-").replace(")", "").replace("Dr.", "").replace("Cr.", "")
    try:
        return Decimal(text)
    except (ValueError, ArithmeticError):
        raise ValueError(f"Non-numeric balance: {value!r}")


class Command(BaseCommand):
    help = "Post opening balances from a TB-style workbook as balanced per-segment JEs."

    def add_arguments(self, parser):
        parser.add_argument("--file", dest="file", default=None, help="CSV/XLSX path")
        parser.add_argument("--company", dest="company", default="STMIET")
        parser.add_argument("--as-of", dest="as_of", default="2026-01-01")
        parser.add_argument("--entry-prefix", dest="prefix", default="OB-2026",
                            help="Entry number prefix, e.g. OB-2026 -> OB-2026-DHPP")
        parser.add_argument("--dr-col", dest="dr_col", default=None,
                            help="Exact header name of the debit balance column")
        parser.add_argument("--cr-col", dest="cr_col", default=None,
                            help="Exact header name of the credit balance column")
        parser.add_argument("--no-plug", action="store_true",
                            help="Do not auto-plug the imbalance; fail instead")

    def _load_rows(self, file_path, dr_header, cr_header):
        if file_path.suffix.lower() == ".csv":
            with open(file_path, newline="", encoding="utf-8-sig") as fh:
                reader = csv.reader(fh)
                header = [h.strip().upper() for h in next(reader, [])]
                for i, row in enumerate(reader, start=2):
                    if not any(row):
                        continue
                    yield self._map_row(row, header, dr_header, cr_header), f"CSV row {i}"
        else:
            if openpyxl is None:
                raise CommandError("openpyxl required for xlsx input; pip install openpyxl")
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header = None
                header_idx = 0
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if not row or not any(row):
                        continue
                    upper = [str(v or "").strip().upper() for v in row]
                    if header is None and any("COA" in h for h in upper):
                        header = upper
                        header_idx = i
                        break
                if header is None:
                    self.stdout.write(self.style.WARNING(
                        f"skip sheet {sheet_name}: no COA header"
                    ))
                    continue
                for i, row in enumerate(ws.iter_rows(min_row=header_idx + 1, values_only=True), start=header_idx + 1):
                    if not row or not any(row):
                        continue
                    yield self._map_row(row, header, dr_header, cr_header), f"{sheet_name} row {i}"

    def _map_row(self, row, header, dr_header, cr_header):
        idx = {}
        for field, names in HEADER_MAP.items():
            for n in names:
                if n in header:
                    idx[field] = header.index(n)
                    break
        if dr_header and dr_header.upper() in header:
            idx["dr"] = header.index(dr_header.upper())
        if cr_header and cr_header.upper() in header:
            idx["cr"] = header.index(cr_header.upper())
        out = {}
        for field, pos in idx.items():
            if pos < len(row):
                out[field] = _clean(row[pos])
        return out

    def _resolve_segment(self, raw, account, company):
        if raw:
            key = raw.upper().strip()
            if key in SINGLE_SEGMENTS or key == "ALL":
                seg = Segment.objects.filter(code=key).first()
                if seg is None and key in SINGLE_SEGMENTS:
                    raise CommandError(
                        f"Segment '{key}' not found. Run import_coa first."
                    )
                return seg
            for alias in SINGLE_SEGMENTS:
                if alias in key:
                    seg = Segment.objects.filter(code=alias).first()
                    if seg is None:
                        raise CommandError(
                            f"Segment '{alias}' not found. Run import_coa first."
                        )
                    return seg
        # Fallback to account tag.
        segment_code = account.segment if account.segment in SINGLE_SEGMENTS else SHARED
        seg = Segment.objects.filter(code=segment_code).first()
        if seg is None and segment_code in SINGLE_SEGMENTS:
            raise CommandError(
                f"Segment '{segment_code}' not found for account {account.code}. "
                "Run import_coa first."
            )
        return seg

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = Path(options["file"]) if options["file"] else None
        if file_path is None or not file_path.exists():
            repo_root = Path(__file__).resolve().parents[5]
            for cand in (
                repo_root / "excel-files" / "OPENING-BALANCES.xlsx",
                repo_root / "excel-files" / "OPENING-BALANCES.csv",
                Path.cwd() / "excel-files" / "OPENING-BALANCES.xlsx",
            ):
                if cand.exists():
                    file_path = cand
                    break
        if file_path is None or not file_path.exists():
            raise CommandError(
                "Opening balances not found. Pass --file, or place "
                "OPENING-BALANCES.csv/.xlsx under /excel-files."
            )

        company = Company.objects.filter(code=options["company"]).first()
        if company is None:
            raise CommandError(f"Company '{options['company']}' not found. Run import_coa first.")

        as_of = date.fromisoformat(options["as_of"])
        prefix = options["prefix"]
        dr_header = options.get("dr_col") or "OPENING DR"
        cr_header = options.get("cr_col") or "OPENING CR"

        # Parse the source into per-segment {account_code: (dr, cr)} maps.
        segment_dr: dict[str, dict[str, Decimal]] = {}
        segment_cr: dict[str, dict[str, Decimal]] = {}
        skipped = 0
        for row, label in self._load_rows(file_path, dr_header, cr_header):
            code = row.get("code", "")
            if not code or not code.isdigit():
                skipped += 1
                continue
            account = Account.objects.filter(code=code, is_postable=True).first()
            if account is None:
                self.stdout.write(self.style.WARNING(
                    f"  {label}: COA account {code} missing in chart — skipped."
                ))
                skipped += 1
                continue
            try:
                dr = _to_decimal(row.get("dr", ""))
                cr = _to_decimal(row.get("cr", ""))
                if not dr and not cr:
                    skipped += 1
                    continue
                segment = self._resolve_segment(row.get("segment", ""), account, company)
                seg_key = segment.code if segment else SHARED
                bucket_dr = segment_dr.setdefault(seg_key, {})
                bucket_cr = segment_cr.setdefault(seg_key, {})
                if dr:
                    bucket_dr[code] = bucket_dr.get(code, Decimal("0.00")) + dr
                if cr:
                    bucket_cr[code] = bucket_cr.get(code, Decimal("0.00")) + cr
            except (ValueError, CommandError) as exc:
                self.stdout.write(self.style.ERROR(f"  {label}: {exc}"))
                skipped += 1

        all_segs = set(segment_dr.keys()) | set(segment_cr.keys())
        if not all_segs:
            raise CommandError("No opening balances parsed from the source file.")

        # Fiscal period for as_of.
        fp = FiscalPeriod.objects.filter(
            fiscal_year__company=company,
            start_date__lte=as_of, end_date__gte=as_of,
        ).first()

        # Post one JE per segment, plus one company-level JE for shared/ALL
        # accounts (e.g. the cash bank accounts) which have no dedicated
        # Segment row. Shared accounts use the company's default segment for
        # the JE header; the GL still carries the entry segment, so per-segment
        # reports see the shared balances under that segment (aggregate SFP/TB
        # sums across segments and therefore stays correct — no double count).
        total_posted_dr = Decimal("0.00")
        total_posted_cr = Decimal("0.00")
        posted_segments = 0
        default_seg = (
            company.segments.order_by("code").first()
            or Segment.objects.filter(company=company).order_by("code").first()
        )
        for seg_key in sorted(all_segs):
            if seg_key == SHARED:
                seg = default_seg
                entry_key = "ALL"
                description_seg = "Shared"
            else:
                seg = Segment.objects.filter(code=seg_key).first()
                entry_key = seg_key
                description_seg = seg_key
            if seg is None:
                self.stdout.write(self.style.WARNING(
                    f"Segment '{seg_key}' not found and no company default — skipping balances for it."
                ))
                continue
            entry_no = f"{prefix}-{entry_key}"
            if JournalEntry.objects.filter(entry_no=entry_no).exists():
                self.stdout.write(self.style.WARNING(
                    f"Entry {entry_no} already exists — skipping (idempotent)."
                ))
                continue

            dr_map = segment_dr.get(seg_key, {})
            cr_map = segment_cr.get(seg_key, {})

            entry = JournalEntry.objects.create(
                entry_no=entry_no,
                company=company,
                segment=seg,
                fiscal_period=fp,
                transaction_date=as_of,
                status=PostingStatus.APPROVED,  # opening > threshold; approved gate
                description=f"Opening balances {description_seg} as of {as_of}",
                source_doc_type="OB",
                source_doc_no=str(as_of),
            )

            line_no = 0
            total_dr = Decimal("0.00")
            total_cr = Decimal("0.00")

            for code in sorted(dr_map):
                amount = dr_map[code]
                if not amount:
                    continue
                line_no += 1
                account = Account.objects.get(code=code)
                JournalEntryLine.objects.create(
                    entry=entry, line_no=line_no, account=account,
                    description=f"Opening {code}",
                    debit=money(amount),
                )
                total_dr += amount

            for code in sorted(cr_map):
                amount = cr_map[code]
                if not amount:
                    continue
                line_no += 1
                account = Account.objects.get(code=code)
                JournalEntryLine.objects.create(
                    entry=entry, line_no=line_no, account=account,
                    description=f"Opening {code}",
                    credit=money(amount),
                )
                total_cr += amount

            # Plug imbalance to the segment's opening equity (segment map).
            difference = total_dr - total_cr
            if difference:
                if options.get("no_plug"):
                    entry.lines.all().delete()
                    entry.delete()
                    raise CommandError(
                        f"Opening balances for {seg_key} out of balance by "
                        f"{money(difference)}. Remove --no-plug to auto-plug."
                    )
                try:
                    plug_account = resolve_segment_account(seg, "opening_equity")
                except AccountingError:
                    self.stdout.write(self.style.ERROR(
                        f"  {seg_key}: no opening_equity map — skipping plug."
                    ))
                    plug_account = None
                if plug_account is not None:
                    line_no += 1
                    if difference > 0:
                        JournalEntryLine.objects.create(
                            entry=entry, line_no=line_no, account=plug_account,
                            description="Opening plug (Dr excess)",
                            credit=money(difference),
                        )
                        total_cr += difference
                    else:
                        JournalEntryLine.objects.create(
                            entry=entry, line_no=line_no, account=plug_account,
                            description="Opening plug (Cr excess)",
                            debit=money(-difference),
                        )
                        total_dr += -difference
                    self.stdout.write(self.style.WARNING(
                        f"  {seg_key}: plugged {money(difference)} to "
                        f"{plug_account.code} {plug_account.name}"
                    ))

            entry.recalc_totals()
            try:
                PostingService.post(entry)
            except AccountingError as exc:
                entry.delete()
                raise CommandError(f"Failed to post opening balances for {seg_key}: {exc}")

            total_posted_dr += total_dr
            total_posted_cr += total_cr
            posted_segments += 1
            self.stdout.write(self.style.SUCCESS(
                f"  posted {entry_no}: Dr {money(total_dr)} = Cr {money(total_cr)}"
            ))

        self.stdout.write(self.style.SUCCESS(
            f"Opening balances posted for {posted_segments} segment(s): "
            f"Dr {money(total_posted_dr)} = Cr {money(total_posted_cr)}. "
            f"{skipped} rows skipped."
        ))
